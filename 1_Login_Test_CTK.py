from PIL import Image
import customtkinter as ctk
from customtkinter import *
from subprocess import call
from tkinter import messagebox
from openpyxl import load_workbook


window = CTk()
window.title("Login Portal")
window.geometry('1300x825')
window.resizable(False, False)

# ==================== Functions ====================
# Login Checker Function
def login_checker():
    if not log_in_data_validation_debugger():
        return

    username = login_user_entry.get()
    password = login_password_entry.get()

    try:
        wb = load_workbook("user_account_data.xlsx")
        ws = wb["Userdata"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                messagebox.showinfo("Login Success", f"Welcome, {username}!")
                return
            
        messagebox.showerror(title= "Login Failed", message= "Incorrect username or password.")

    except FileNotFoundError:
        messagebox.showerror(title= "Error", message= "User data file not found.")

# Input Validation & Debugger Function
def log_in_data_validation_debugger():

    username = login_user_entry.get()
    password = login_password_entry.get()

    if username and password and username != "Username" and password != "Password":
        print("\n\nData Entry Form:\n\n" \
              "Username:",username,
              "\nPassword:",password,
              "\n")
        return True
    
    else:
        messagebox.showerror(title= "Error", message= "Error. Input Required")
        return False

# Username Delete & Restore Function
def log_in_on_username_click(event):
    if login_user_entry.get() == "Username":
        login_user_entry.delete(0, 'end')

def log_in_on_username_leave(event):
    name = login_user_entry.get()
    if name == "":
        login_user_entry.insert(0, "Username")

# Password Delete & Restore Function
def log_in_on_password_click(event):
    if login_password_entry.get() == "Password":
        login_password_entry.delete(0, END)

def log_in_on_password_leave(event):
    password = login_password_entry.get()
    if password == "":
        login_password_entry.insert(0, "Password")

# Signup Window Switch Function
def go_signup():
    window.destroy()
    call(["python", "2_Sign_Up_Test_CTK.py"])

def log_in_handle_enter(event):
    if not login_checker():
        return
    login_button.invoke()

# //////////////////////////////////////////////////////////

# ==================== UI ====================
# Background & Banner Img
background_image = Image.open(".\wallhaven-85gxp2.png")
bg_img = CTkImage(light_image=background_image, dark_image=background_image, size=(1300, 825))
bg_label = CTkLabel(window, image=bg_img, text="")
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

side_image = Image.open(".\wallhaven-73616y.png")
side_img = CTkImage(light_image=side_image, dark_image=side_image, size=(550, 550))
side_label = CTkLabel(window, image=side_img, text="", corner_radius=10)
side_label.place(relx = 0.5, rely = 0.5, x= -547, y= -268)

# //////////////////////////////////////////////////////////

# ==================== Frames ====================
log_in_frame = CTkFrame(window, border_width= 3, corner_radius= 15)
log_in_frame.place(relx = 0.5, rely = 0.5, x= 90, y= -200)

# //////////////////////////////////////////////////////////

# ==================== Labels & Inputs ====================
# Project Name
login_project_label = CTkLabel(log_in_frame, text= "PROJECT UniPass", font= ("Times New Roman bold", 40))
login_project_label.grid(row=0, column=0, sticky="n", pady= 10, padx= 15)

# Short Description
short_desc_label =  CTkLabel(log_in_frame, text= "/Short Description/", font= ("Helvetica bold", 18))
short_desc_label.grid(row=1, column=0, sticky="n")

# Login
login_label = CTkLabel(log_in_frame, text= "Log In to Project", font= ("Helvetica bold", 17))
login_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)

# User Entry
login_user_entry = CTkEntry(log_in_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Username", height= 35)
login_user_entry.grid(row=3, column=0, sticky= "n", pady=15, padx= 15)
login_user_entry.bind("<FocusIn>", log_in_on_username_click)
login_user_entry.bind("<FocusOut>", log_in_on_username_leave)
login_user_entry.bind('<Return>', log_in_handle_enter)

# Divider Line
user_line = CTkFrame(log_in_frame, width=400, height=2, fg_color="white")
user_line.place(x=15, y=203)

# Password Entry
login_password_entry = CTkEntry(log_in_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Password", show="*", height= 35)
login_password_entry.grid(row=4, column=0, sticky= "n", pady=15, padx= 15)
login_password_entry.bind("<FocusIn>", log_in_on_password_click)
login_password_entry.bind("<FocusOut>", log_in_on_password_leave)
login_password_entry.bind('<Return>', log_in_handle_enter)

# Divider Line
pass_line = CTkFrame(log_in_frame, width=400, height=2, fg_color="white")
pass_line.place(x=15, y=268)

# //////////////////////////////////////////////////////////

# ==================== Buttons ====================

# Button Login
login_button = CTkButton(log_in_frame, text= "Log In", width=325, font= ("Arial bold", 15), command= login_checker, height= 35)
login_button.grid(row=5, column=0, pady=15, padx= 15)

# Signup Text + Button
sign_frame = CTkFrame(log_in_frame)
sign_frame.grid(row=6, column=0, pady=20)

need_account_label = CTkLabel(sign_frame, text="Need an Account?", font=("Arial", 12))
need_account_label.grid(row=0, column=0, padx=10)

sign_up_button = CTkButton(sign_frame, text="SIGN UP", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2", command= go_signup, width=50)
sign_up_button.grid(row=0, column=1, padx=5)

# //////////////////////////////////////////////////////////

# ==================== Window Starter ====================
window.mainloop()