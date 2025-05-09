from customtkinter import *
import customtkinter as ctk
from datetime import datetime
from tkinter import filedialog
from PIL import Image
from tkinter import messagebox
import subprocess
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re
import random
    
# ==================== Window Setup ====================
window = CTk()
window.title("Student Enrollment Form")
height = 825
width = 1300
x = (window.winfo_screenwidth()//2)-(width//2) 
y = (window.winfo_screenheight()//2)-(height//2) 
window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
window.resizable(False, False)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ==================== Functions ====================
# Creating a new sheet in the workbook and saving it
def create_new_sheet():
    if not FinalCheck():
        return
    
    try:

        sections = Course_Section_Entry.get().strip()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror(title="Error", message="Please select a valid section.")
            return
        
        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        try:
            wb = load_workbook(file_path_to_excel)
        except FileNotFoundError:
            wb = Workbook()
            wb.save(file_path_to_excel)

        if new_sheet_name in wb.sheetnames:
            new_sheet = wb[new_sheet_name]    

        else:
            new_sheet = wb.create_sheet(new_sheet_name)
            new_headers = [
                    "Student ID", "Course/Section", "LRN", 
                    "", 
                    "Surname", "Firstname", "Middle Initial", 
                    "", 
                    "Gender", "Age", "Birthdate", "Birthplace", "Nationality", "Religion", "Marital Status", "Language Spoken", 
                    "", 
                    "Street", "Barangay", "City", "Zip Code", "Province", "Country", 
                    "", 
                    "Email", "Contact Number",
                    "",
                    "Father's Name", "Father's Address", "Father's Occupation", "Father's Contact No",
                    "",
                    "Mother's Name", "Mother's Address","Mother's Occupation", "Mother's Contact No",
                    "",
                    "Guardian's Name", "Guardian's Relationship", "Guardian's Address", "Guardian's Occupation", "Guardian's Contact No",
                    "",
                    "Elementary School", "Elementary Address", "Elementary Year Graduated",
                    "",
                    "Junior High School", "Junior High Address", "Junior High Year Graduated",
                    "",
                    "Senior High School", "Senior High Address", "Senior High Strand", "Senior High Year Graduated",
                    "",
                    "College", "College Address", "College Year Graduated"
                        ]
            new_sheet.append(new_headers)

        student_info_data = [
                Student_Entry.get().strip(), Course_Section_Entry.get().strip(), LRN_Entry.get().strip(), 
                ""
                ]
        student_name_data = [
            Surname_Entry.get().strip().strip(), FirstName_Entry.get().strip(), Middle_Entry.get().strip(), 
            ""
                ]
        
        student_personal_detail_data = [
                "Male" if gender_var.get() == 1 else "Female" if gender_var.get() == 2 else "Prefer not to answer", 
                age_box.get().strip(), f"{month_box.get().strip()} {day_box.get().strip()}, {year_box.get().strip()}",
                Birthplace_Entry.get().strip(), Nationality_Entry.get().strip(), religion_box.get().strip(), marital_status_box.get().strip(), 
                Language_Entry.get().strip(), 
                "", 
                Street_Entry.get().strip(), BRGY_Entry.get().strip(), City_Entry.get().strip(),
                ZipCode_Entry.get().strip(), Province_Entry.get().strip(), Country_Entry.get().strip(), 
                "",
                Email_Entry.get().strip(), Num_Entry.get().strip(),
                ""
                ]
        student_family_detail_data = [
                Name_Father_Entry.get().strip(), Address_Father_Entry.get().strip(), Occupation_Father_Entry.get().strip(), Phone_Father_Entry.get().strip(),
                "",
                Name_Mother_Entry.get().strip(), Address_Mother_Entry.get().strip(), Occupation_Mother_Entry.get().strip(), Phone_Mother_Entry.get().strip(),
                "",
                Name_Guardian_Entry.get().strip(), Rel_Guardian_Entry.get().strip(), Address_Guardian_Entry.get().strip(),
                Occupation_Guardian_Entry.get().strip(), Phone_Guardian_Entry.get().strip(),
                ""
                ]
        student_educational_detail_data = [
                Schl_Elem_Entry.get().strip(), Address_Elem_Entry.get().strip(), YR_Elem_Entry.get().strip(),
                "",
                Schl_Junior_Entry.get().strip(), Address_Junior_Entry.get().strip(), YR_Junior_Entry.get().strip(),
                "",
                Schl_Senior_Entry.get().strip(), Address_Senior_Entry.get().strip(), strand_Senior_Entry.get().strip(), YR_Senior_Entry.get().strip(),
                "",
                Schl_College_Entry.get().strip(), Address_College_Entry.get().strip(), YR_College_Entry.get().strip()
                ]
        
        new_sheet.append(student_info_data + student_name_data + student_personal_detail_data + student_family_detail_data + student_educational_detail_data)
        

        wb.save(file_path_to_excel)
        messagebox.showinfo(title= "Success", message= "Data saved successfully!")
        format_excel()

        fields_to_readonly = [
                Student_Entry, generate_student_id, Course_Section_Entry, LRN_Entry, generate_student_lrn,
                Surname_Entry, FirstName_Entry, Middle_Entry, male_checkbox, female_checkbox, none_binary_checkbox, 
                age_box, month_box, day_box, year_box, Birthplace_Entry, Nationality_Entry, religion_box, 
                marital_status_box, Language_Entry, Street_Entry, BRGY_Entry, City_Entry, ZipCode_Entry, 
                Province_Entry, Country_Entry, Email_Entry, Num_Entry, Name_Father_Entry, Address_Father_Entry, 
                Occupation_Father_Entry, Phone_Father_Entry, Name_Mother_Entry, Address_Mother_Entry, Occupation_Mother_Entry, 
                Phone_Mother_Entry, Name_Guardian_Entry, Rel_Guardian_Entry, Address_Guardian_Entry,
                Occupation_Guardian_Entry, Phone_Guardian_Entry, same_chk_box, Schl_Elem_Entry, Address_Elem_Entry,
                YR_Elem_Entry, Schl_Junior_Entry, Address_Junior_Entry, YR_Junior_Entry, Schl_Senior_Entry, Address_Senior_Entry,
                strand_Senior_Entry, YR_Senior_Entry, Schl_College_Entry, Address_College_Entry, YR_College_Entry, if_transferee_chk_box
        ]
        for field in fields_to_readonly:
            try:
                field.configure(state="readonly")
            except AttributeError:
                pass

        gender_var.set(0)
        age_box.set("Select Age")
        month_box.set("MM")
        day_box.set("DD")
        year_box.set("YYYY")
        religion_box.set("Select Religion")
        marital_status_box.set("Select Marital Status")

    except Exception as e:
        messagebox.showerror(title= "Error", message= f"An error occurred: {e}")

# Validating the User per Email
def validating_user_email():
    try:

        email = Email_Entry.get().strip()
        if not email:
            messagebox.showerror(title="Error", message="Please enter an email Address.")
            return

        wb = load_workbook("user_account_data.xlsx")
        account_data_sheet = wb["Userdata"]
        for row in account_data_sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == email:
                Surname_Entry.delete(0, END)
                Surname_Entry.insert(0, row[1])

                FirstName_Entry.delete(0, END)
                FirstName_Entry.insert(0, row[0])

                Email_Entry.delete(0, END)
                Email_Entry.insert(0, row[2])
                return 
        
        messagebox.showerror(title="Error", message="User not found.")
            
    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")

# Generating a Formatted User ID with Random Numbers
def generate_user_id(YR: int = 2025, existing_ids=None):
    try:

        section = Course_Section_Entry.get().strip()

        if section not in ["1A", "1B", "1C"]:
            raise ValueError("Section must be '1A', '1B', or '1C'.")

        year_part = str(YR)[1:]

        if existing_ids is None:
            existing_ids = set()

        while True:
            number_part = f"{random.randint(0, 9999):04}"
            user_id = f"{year_part}{section}-{number_part}"

            if user_id not in existing_ids:
                Student_Entry.delete(0, END)
                Student_Entry.insert(0, user_id)
                Student_Entry.configure(state="readonly")
                Course_Section_Entry.configure(state="readonly")
                return user_id

    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")

# Generating a Formatted User LRN with Random Numbers
def generate_user_lrn(special_num: int = 20011213, existing_lrns=None):
    try:

        format_lrn = str(special_num)

        if existing_lrns is None:
            existing_lrns = set()

        while True:
            number_part = f"{random.randint(0, 9999):04}"
            user_lrn = f"{format_lrn}{number_part}"

            if user_lrn not in existing_lrns:
                LRN_Entry.delete(0, END)
                LRN_Entry.insert(0, user_lrn)
                LRN_Entry.configure(state="readonly")
                return user_lrn

    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")

# Validating the User per Password
def ask_password():
    result = {"success": False}

    def check_password():
        try:
            file_path_to_excel = "user_account_data.xlsx"
            wb = load_workbook(file_path_to_excel)
            account_ask_password = wb["Userdata"]

            for row in account_ask_password.iter_rows(min_row=2, values_only=True):
                if row[4] == Password_Entry.get().strip():
                    result["success"] = True
                    ask_password_window.destroy()
                    return
                
            messagebox.showerror(title= "Access Denied", message= "Incorrect Password.")

        except FileNotFoundError:
            messagebox.showerror(title="Error", message="The file 'user_account_data.xlsx' was not found.")
        except Exception as e:
            messagebox.showerror(title="Error", message=f"An error occurred: {e}")

    ask_password_window = CTkToplevel(window)
    ask_password_window.title("Password Required")
    height = 160
    width = 300
    x = (ask_password_window.winfo_screenwidth()//2)-(width//2) 
    y = (ask_password_window.winfo_screenheight()//2)-(height//2) 
    ask_password_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    ask_password_window.resizable(False, False)
    ask_password_window.grab_set()

    pass_label = CTkLabel(ask_password_window, text="Enter Password:", font=("Arial", 16))
    pass_label.pack(pady=10)

    Password_Entry = CTkEntry(ask_password_window, show="*", width=200)
    Password_Entry.pack(pady=5)
    Password_Entry.focus()

    pass_submit_btn = ctk.CTkButton(ask_password_window, text="Submit", command=check_password)
    pass_submit_btn.pack(pady=15)

    ask_password_window.wait_window()
    return result["success"]

# Loading the user's data
def load_user_data():
    try:

        student_user_account = Student_Entry.get().strip()
        student_sections = Course_Section_Entry.get().strip()
        student_lrn = LRN_Entry.get().strip()

        if not student_user_account:
            messagebox.showerror(title="Error", message="Please enter your Student ID.")
            return
        
        if student_sections not in ["1A", "1B", "1C"]:
            messagebox.showerror(title="Error", message="Please select a valid section.")
            return
        
        if not student_lrn:
            messagebox.showerror(title="Error", message="Please enter your LRN.")
            return

        new_sheet_name = f"{student_sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror(title="Error", message=f"Sheet '{new_sheet_name}' does not exist.")
            return
        
        student_details_sheet = wb[new_sheet_name]

        for row in student_details_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_user_account and row[1] == student_sections and row[2] == student_lrn:

                if not ask_password():
                    return

                Surname_Entry.delete(0, END)
                Surname_Entry.insert(0, row[4])
                FirstName_Entry.delete(0, END)
                FirstName_Entry.insert(0, row[5])
                Middle_Entry.delete(0, END)
                Middle_Entry.insert(0, row[6])

                gender = row[8]
                if gender == "Male":
                    gender_var.set(1)
                elif gender == "Female":
                    gender_var.set(2)
                elif gender == "Prefer not to answer":
                    gender_var.set(3)
                else:
                    gender_var.set(0)
                age_box.set(row[9])
                birthdate = row[10]
                if birthdate:
                    try:
                        month, day_year = birthdate.split(" ", 1)
                        day, year = day_year.replace(",", "").split(" ")
                        month_box.set(month)
                        day_box.set(day)
                        year_box.set(year)
                    except ValueError:
                        month_box.set("MM")
                        day_box.set("DD")
                        year_box.set("YYYY")
                else:
                    month_box.set("MM")
                    day_box.set("DD")
                    year_box.set("YYYY")
                Birthplace_Entry.delete(0, END)
                Birthplace_Entry.insert(0, row[11])
                Nationality_Entry.delete(0, END)
                Nationality_Entry.insert(0, row[12])
                religion_box.set(row[13])
                marital_status_box.set(row[14])
                Language_Entry.delete(0, END)
                Language_Entry.insert(0, row[15])

                Street_Entry.delete(0, END)
                Street_Entry.insert(0, row[17])
                BRGY_Entry.delete(0, END)
                BRGY_Entry.insert(0, row[18])
                City_Entry.delete(0, END)
                City_Entry.insert(0, row[19])
                ZipCode_Entry.delete(0, END)
                ZipCode_Entry.insert(0, row[20])
                Province_Entry.delete(0, END)
                Province_Entry.insert(0, row[21])
                Country_Entry.delete(0, END)
                Country_Entry.insert(0, row[22])

                Email_Entry.delete(0, END)
                Email_Entry.insert(0, row[24])
                Num_Entry.delete(0, END)
                Num_Entry.insert(0, row[25])

                Name_Father_Entry.delete(0, END)
                Name_Father_Entry.insert(0, row[27])
                Address_Father_Entry.delete(0, END)
                if row[28] is not None and isinstance(row[28], str):
                    Address_Father_Entry.insert(0, row[28])
                elif row[28] is not None:
                    Address_Father_Entry.insert(0, str(row[28]))
                else:
                    Address_Father_Entry.insert(0, "")
                Occupation_Father_Entry.delete(0, END)
                Occupation_Father_Entry.insert(0, row[29])                
                Phone_Father_Entry.delete(0, END)
                Phone_Father_Entry.insert(0, row[30])

                Name_Mother_Entry.delete(0, END)
                Name_Mother_Entry.insert(0, row[32])
                Address_Mother_Entry.delete(0, END)
                Address_Mother_Entry.insert(0, row[33])
                Occupation_Mother_Entry.delete(0, END)
                Occupation_Mother_Entry.insert(0, row[34])                
                Phone_Mother_Entry.delete(0, END)
                Phone_Mother_Entry.insert(0, row[35])

                Name_Guardian_Entry.delete(0, END)
                if row[37] is not None and isinstance(row[37], str):
                    Name_Guardian_Entry.insert(0, row[37])
                elif row[37] is not None:
                    Name_Guardian_Entry.insert(0, str(row[37]))
                else:
                    Name_Guardian_Entry.insert(0, "")
                Rel_Guardian_Entry.delete(0, END)
                if row[38] is not None and isinstance(row[38], str):
                    Rel_Guardian_Entry.insert(0, row[38])
                elif row[38] is not None:
                    Rel_Guardian_Entry.insert(0, str(row[38]))
                else:
                    Rel_Guardian_Entry.insert(0, "")
                Address_Guardian_Entry.delete(0, END)
                if row[39] is not None and isinstance(row[39], str):
                    Address_Guardian_Entry.insert(0, row[39])
                elif row[39] is not None:
                    Address_Guardian_Entry.insert(0, str(row[39]))
                else:
                    Address_Guardian_Entry.insert(0, "")
                Occupation_Guardian_Entry.delete(0, END)
                if row[40] is not None and isinstance(row[40], str):
                    Occupation_Guardian_Entry.insert(0, row[40])
                elif row[40] is not None:
                    Occupation_Guardian_Entry.insert(0, str(row[40]))
                else:
                    Occupation_Guardian_Entry.insert(0, "")                                           
                Phone_Guardian_Entry.delete(0, END)
                if row[41] is not None and isinstance(row[41], str):
                    Phone_Guardian_Entry.insert(0, row[41])
                elif row[41] is not None:
                    Phone_Guardian_Entry.insert(0, str(row[41]))
                else:
                    Phone_Guardian_Entry.insert(0, "")

                Schl_Elem_Entry.delete(0, END)
                Schl_Elem_Entry.insert(0, row[43])
                Address_Elem_Entry.delete(0, END)
                Address_Elem_Entry.insert(0, row[44])
                YR_Elem_Entry.delete(0, END)
                YR_Elem_Entry.insert(0, row[45])

                Schl_Junior_Entry.delete(0, END)
                Schl_Junior_Entry.insert(0, row[47])
                Address_Junior_Entry.delete(0, END)
                Address_Junior_Entry.insert(0, row[48])
                YR_Junior_Entry.delete(0, END)
                YR_Junior_Entry.insert(0, row[49])

                Schl_Senior_Entry.delete(0, END)
                Schl_Senior_Entry.insert(0, row[51])
                Address_Senior_Entry.delete(0, END)
                Address_Senior_Entry.insert(0, row[52])
                strand_Senior_Entry.delete(0, END)
                strand_Senior_Entry.insert(0, row[53])
                YR_Senior_Entry.delete(0, END)
                YR_Senior_Entry.insert(0, row[54])

                Schl_College_Entry.delete(0, END)
                if row[56] is not None and isinstance(row[56], str):
                    Schl_College_Entry.insert(0, row[56])
                elif row[56] is not None:
                    Schl_College_Entry.insert(0, str(row[56]))
                else:
                    Schl_College_Entry.insert(0, "")
                Address_College_Entry.delete(0, END)
                if row[57] is not None and isinstance(row[57], str):
                    Address_College_Entry.insert(0, row[57])
                elif row[57] is not None:
                    Address_College_Entry.insert(0, str(row[57]))
                else:
                    Address_College_Entry.insert(0, "")
                YR_College_Entry.delete(0, END)
                if row[58] is not None and isinstance(row[58], str):
                    YR_College_Entry.insert(0, row[58])
                elif row[58] is not None:
                    YR_College_Entry.insert(0, str(row[58]))
                else:
                    YR_College_Entry.insert(0, "")

                fields_to_disable = [
                        Student_Entry, generate_student_id, Course_Section_Entry, LRN_Entry, generate_student_lrn,
                        Surname_Entry, FirstName_Entry, Middle_Entry, male_checkbox, female_checkbox, none_binary_checkbox, 
                        age_box, month_box, day_box, year_box, Birthplace_Entry, Nationality_Entry, religion_box, 
                        marital_status_box, Language_Entry, Street_Entry, BRGY_Entry, City_Entry, ZipCode_Entry, 
                        Province_Entry, Country_Entry, Email_Entry, Num_Entry, Name_Father_Entry, Address_Father_Entry, 
                        Occupation_Father_Entry, Phone_Father_Entry, Name_Mother_Entry, Address_Mother_Entry, Occupation_Mother_Entry, 
                        Phone_Mother_Entry, Name_Guardian_Entry, Rel_Guardian_Entry, Address_Guardian_Entry,
                        Occupation_Guardian_Entry, Phone_Guardian_Entry, same_chk_box, Schl_Elem_Entry, Address_Elem_Entry,
                        YR_Elem_Entry, Schl_Junior_Entry, Address_Junior_Entry, YR_Junior_Entry, Schl_Senior_Entry, Address_Senior_Entry,
                        strand_Senior_Entry, YR_Senior_Entry, Schl_College_Entry, Address_College_Entry, YR_College_Entry, if_transferee_chk_box
                ]
                for field in fields_to_disable:
                    try:
                        field.configure(state="disable")
                    except AttributeError:
                        pass

                return

        messagebox.showerror(title="Error", message="User not found.")
            
    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")

def edit_user_data():
    if not ask_password():
        return
        
    fields_to_enable = [
            Student_Entry, generate_student_id, Course_Section_Entry, LRN_Entry, generate_student_lrn,
            Surname_Entry, FirstName_Entry, Middle_Entry, male_checkbox, female_checkbox, none_binary_checkbox, 
            age_box, month_box, day_box, year_box, Birthplace_Entry, Nationality_Entry, religion_box, 
            marital_status_box, Language_Entry, Street_Entry, BRGY_Entry, City_Entry, ZipCode_Entry, 
            Province_Entry, Country_Entry, Email_Entry, Num_Entry, Name_Father_Entry, Address_Father_Entry, 
            Occupation_Father_Entry, Phone_Father_Entry, Name_Mother_Entry, Address_Mother_Entry, Occupation_Mother_Entry, 
            Phone_Mother_Entry, Name_Guardian_Entry, Rel_Guardian_Entry, Address_Guardian_Entry,
            Occupation_Guardian_Entry, Phone_Guardian_Entry, same_chk_box, Schl_Elem_Entry, Address_Elem_Entry,
            YR_Elem_Entry, Schl_Junior_Entry, Address_Junior_Entry, YR_Junior_Entry, Schl_Senior_Entry, Address_Senior_Entry,
            strand_Senior_Entry, YR_Senior_Entry, Schl_College_Entry, Address_College_Entry, YR_College_Entry, if_transferee_chk_box
    ]
    for field in fields_to_enable:
        try:
            field.configure(state="normal")
        except AttributeError:
            pass

def update_user_data():
    try:
        student_user_account = Student_Entry.get().strip()
        student_sections = Course_Section_Entry.get().strip()
        student_lrn = LRN_Entry.get().strip()

        if not student_user_account:
            messagebox.showerror(title="Error", message="Please enter your Student ID.")
            return
        
        if student_sections not in ["1A", "1B", "1C"]:
            messagebox.showerror(title="Error", message="Please select a valid section.")
            return
        
        if not student_lrn:
            messagebox.showerror(title="Error", message="Please enter your LRN.")
            return

        new_sheet_name = f"{student_sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror(title="Error", message=f"Sheet '{new_sheet_name}' does not exist.")
            return

        ws = wb[new_sheet_name]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == Student_Entry.get().strip() and row[1].value == Course_Section_Entry.get().strip() and row[2].value == LRN_Entry.get().strip():

                if not ask_password():
                    return

                row[4].value = Surname_Entry.get().strip()
                row[5].value = FirstName_Entry.get().strip()
                row[6].value = Middle_Entry.get().strip()

                row[8].value = "Male" if gender_var.get() == 1 else "Female" if gender_var.get() == 2 else "Prefer not to answer"
                row[9].value = age_box.get().strip()
                row[10].value = f"{month_box.get().strip()} {day_box.get().strip()}, {year_box.get().strip()}"
                row[11].value = Birthplace_Entry.get().strip()
                row[12].value = Nationality_Entry.get().strip()
                row[13].value = religion_box.get().strip()
                row[14].value = marital_status_box.get().strip()
                row[15].value = Language_Entry.get().strip()

                row[17].value = Street_Entry.get().strip()
                row[18].value = BRGY_Entry.get().strip()
                row[19].value = City_Entry.get().strip()
                row[20].value = ZipCode_Entry.get().strip()
                row[21].value = Province_Entry.get().strip()
                row[22].value = Country_Entry.get().strip()

                row[24].value = Email_Entry.get().strip()
                row[25].value = Num_Entry.get().strip()

                row[27].value =  Name_Father_Entry.get().strip()
                row[28].value = Address_Father_Entry.get().strip()
                row[29].value = Occupation_Father_Entry.get().strip()
                row[30].value = Phone_Father_Entry.get().strip()

                row[32].value = Name_Mother_Entry.get().strip()
                row[33].value = Address_Mother_Entry.get().strip()
                row[34].value = Occupation_Mother_Entry.get().strip()
                row[35].value = Phone_Mother_Entry.get().strip()

                row[37].value = Name_Guardian_Entry.get().strip()
                row[38].value = Rel_Guardian_Entry.get().strip()
                row[39].value = Address_Guardian_Entry.get().strip()
                row[40].value = Occupation_Guardian_Entry.get().strip()
                row[41].value = Phone_Guardian_Entry.get().strip()

                row[43].value = Schl_Elem_Entry.get().strip()
                row[44].value = Address_Elem_Entry.get().strip()
                row[45].value = YR_Elem_Entry.get().strip()

                row[47].value = Schl_Junior_Entry.get().strip()
                row[48].value = Address_Junior_Entry.get().strip()
                row[49].value = YR_Junior_Entry.get().strip()

                row[51].value = Schl_Senior_Entry.get().strip()
                row[52].value = Address_Senior_Entry.get().strip()
                row[53].value = strand_Senior_Entry.get().strip()
                row[54].value = YR_Senior_Entry.get().strip()

                row[56].value = Schl_College_Entry.get().strip()
                row[57].value = Address_College_Entry.get().strip()
                row[58].value = YR_College_Entry.get().strip()

                wb.save(file_path_to_excel)
                
                messagebox.showinfo(title="Success", message="User Data Updated Successfully.")
                return
            
            messagebox.showerror(title="Error", message="User not found.") 
    
    except Exception as e:
        messagebox.showerror(title="Error", message=str(e))

# Format Fixer Function
def format_excel():
    try:
        sections = Course_Section_Entry.get().strip()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror(title="Error", message="Invalid section selected.")
            return

        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name not in wb.sheetnames:
            messagebox.showerror(title="Error", message=f"Sheet '{new_sheet_name}' does not exist.")
            return
        
        ws = wb[new_sheet_name]

        for cells in ws[1]:
            if cells.value:
                cells.font = Font(bold=True)

        for cols in ws.columns:
            max_length = 0
            col_letter = get_column_letter(cols[0].column)
            for cell in cols:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path_to_excel)

    except FileNotFoundError:
        messagebox.showerror(title="Error", message="The Excel file was not found.")
    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")


# ==================== Validators ====================
# Final Check Function
def FinalCheck():
    # ==================== PERSONAL DETAILS PAGE ====================
    # Validate Student ID, Course/Section, and LRN fields
    ReqFields = [
        ("Student ID", Student_Entry), 
        ("Course/Section", Course_Section_Entry), 
        ("LRN", LRN_Entry),
    ]
    for Fieldnames, field in ReqFields:
        try:
            if field.get().strip() == "":
                messagebox.showerror(title="Missing Requirement", message=f"Error: {Fieldnames} is required.")
                return False
        except AttributeError:
            messagebox.showerror(title="Internal Error", message=f"Field '{Fieldnames}' is not a valid input widget.")
            return False
        
    selected_section = Course_Section_Entry.get().strip()
    if selected_section == "Select Section":
        messagebox.showerror(title="Missing Requirement", message="Error: Please select a course/section.")
        return False
    
    lrn_value = LRN_Entry.get().strip()
    if not lrn_value.isdigit():
        messagebox.showerror(title="Invalid Input", message="Error: LRN must be a valid integer.")
        return False

    # Validate Full Name fields
    FullName = [
        ("Surname", Surname_Entry), 
        ("Firstname", FirstName_Entry), 
        ("Middle Initial", Middle_Entry),
    ]
    
    for Fieldnames, field in FullName:
        field_value = field.get().strip()
        
        
        if field_value == "" or (field_value == "Surname" and Fieldnames == "Surname") or \
           (field_value == "Firstname" and Fieldnames == "Firstname") or \
           (field_value == "M.I." and Fieldnames == "Middle Initial"):
            messagebox.showerror(title="Missing Requirement", message=f"Error: {Fieldnames} is required.")
            return False
        
        if not field_value.isalpha() and field_value != "":
            messagebox.showerror(title="Invalid Input", message=f"Error: {Fieldnames} should contain only letters.")
            return False

    # Check Gender Section
    if gender_var.get() == 0:
        messagebox.showerror(title="Missing Requirement", message="Error: Please select a gender.")
        return False

    # Validate Age field
    if age_box.get() == "Select Age":
        messagebox.showerror(title="Missing Requirement", message="Error: Please select your age.")
        return False

    # Validate Birthdate (Month, Day, Year)
    if month_box.get() in ["MM", "Select Month"]:
        messagebox.showerror(title="Missing Requirement", message="Error: Please select a month.")
        return False
    if day_box.get() in ["DD", "Select Day"]:
        messagebox.showerror(title="Missing Requirement", message="Error: Please select a day.")
        return False
    if year_box.get() in ["YYYY", "Select Year"]:
        messagebox.showerror(title="Missing Requirement", message="Error: Please select a year.")
        return False

    # Validate Birthplace
    birthplace_value = Birthplace_Entry.get().strip()
    if birthplace_value == "":
        messagebox.showerror(title="Missing Requirement", message="Error: Birthplace is required.")
        return False
    if not birthplace_value.isalpha():
        messagebox.showerror(title="Invalid Input", message="Error: Birthplace should contain only letters.")
        return False

    # Validate Nationality
    nationality_value = Nationality_Entry.get().strip()
    if nationality_value == "":
        messagebox.showerror(title="Missing Requirement", message="Error: Nationality is required.")
        return False
    if not nationality_value.isalpha():
        messagebox.showerror(title="Invalid Input", message="Error: Nationality should contain only letters.")
        return False

    # Validate Religion
    if religion_box.get() == "Select Religion":
        messagebox.showerror(title="Missing Requirement", message="Error: Religion is required.")
        return False

    # Validate Marital Status
    if marital_status_box.get() == "Select Marital Status":
        messagebox.showerror(title="Missing Requirement", message="Error: Marital Status is required.")
        return False

    # Validate Language Spoken
    language_value = Language_Entry.get().strip()
    if language_value == "":
        messagebox.showerror(title="Missing Requirement", message="Error: Language Spoken is required.")
        return False
    if not language_value.isalpha():
        messagebox.showerror(title="Invalid Input", message="Error: Language Spoken should contain only letters.")
        return False

    # Validate Address Section (Street, Barangay, City/Municipality, Province, Zip Code, Country)
    Address = [
        ("Street", Street_Entry),
        ("Barangay", BRGY_Entry),
        ("City/Municipality", City_Entry),
        ("Zip Code", ZipCode_Entry),
        ("Province", Province_Entry),
        ("Country", Country_Entry),
    ]
    for Fieldnames, Entry in Address:
        field_value = Entry.get().strip()
        if not field_value:
            messagebox.showerror(title="Missing Requirement", message=f"Error: {Fieldnames} is required.")
            return False
        if Fieldnames == "Zip Code":
            if not field_value.isdigit() or len(field_value) != 4:
                messagebox.showerror(title="Invalid Input", message=f"Error: {Fieldnames} must be a 4-digit number.")
                return False

        if Fieldnames in ["Street", "Barangay", "City/Municipality", "Province", "Country"]:
            if not field_value.replace(" ", "").isalpha():
                messagebox.showerror(title="Invalid Input", message=f"Error: {Fieldnames} should contain only letters.")
                return False
    
    # Validate Contact Information
    email_value = Email_Entry.get().strip()
    phone_value = Num_Entry.get().strip()

    if email_value == "":
        messagebox.showerror(title="Missing Requirement", message="Error: Email is required.")
        return False
    email_regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    if not re.match(email_regex, email_value):
        messagebox.showerror(title="Invalid Input", message="Error: Please enter a valid email Address.")
        return False

    if phone_value == "":
        messagebox.showerror(title="Missing Requirement", message="Error: Phone Number is required.")
        return False
    if not phone_value.isdigit() and not phone_value.startswith("+"):
        messagebox.showerror(title="Invalid Input", message="Error: Phone Number must be a valid integer or start with '+'.")
        return False

    # ==================== FAMILY BACKGROUND PAGE ====================
    # Validate Parents' Information
    father_name = Name_Father_Entry.get().strip()
    father_occupation = Occupation_Father_Entry.get().strip()
    father_contact = Phone_Father_Entry.get().strip()

    if father_name == "" or father_occupation == "" or father_contact == "":
        messagebox.showerror("Missing Requirement", "Error: Father's information is incomplete.")
        return False
    if not father_name.isalpha() or not father_occupation.isalpha():
        messagebox.showerror("Invalid Input", "Error: Father's Name and Occupation should be strings.")
        return False
    if father_contact != "N/A" and not father_contact.isdigit():
        messagebox.showerror("Invalid Input", "Error: Father's Contact No should be an integer or 'N/A'.")
        return False

    mother_name = Name_Mother_Entry.get().strip()  # Name Entry
    mother_occupation = Occupation_Mother_Entry.get().strip()  # Occupation Entry
    mother_contact = Phone_Mother_Entry.get().strip()  # Contact Entry

    if mother_name == "" or mother_occupation == "" or mother_contact == "":
        messagebox.showerror("Missing Requirement", "Error: Mother's information is incomplete.")
        return False
    if not mother_name.isalpha() or not mother_occupation.isalpha():
        messagebox.showerror("Invalid Input", "Error: Mother's Name and Occupation should be strings.")
        return False
    if mother_contact != "N/A" and not mother_contact.isdigit():
        messagebox.showerror("Invalid Input", "Error: Mother's Contact No should be an integer or 'N/A'.")
        return False

    # Validate Guardian's Information
    if same_chk_var.get() == 0:
        guardian_name = Name_Guardian_Entry.get().strip()
        guardian_relationship = Rel_Guardian_Entry.get().strip()
        guardian_Address = Address_Guardian_Entry.get().strip()
        guardian_occupation = Occupation_Guardian_Entry.get().strip()
        guardian_contact = Phone_Guardian_Entry.get().strip()

        
        if guardian_name == "" or guardian_relationship == "" or guardian_Address == "" or guardian_occupation == "" or guardian_contact == "":
            messagebox.showerror("Missing Requirement", "Error: Guardian's information is incomplete.")
            return False
        if not guardian_name.replace(" ", "").isalpha():
            messagebox.showerror("Invalid Input", "Error: Guardian's Name should contain only letters.")
            return False
        if not guardian_relationship.replace(" ", "").isalpha():
            messagebox.showerror("Invalid Input", "Error: Guardian's Relationship should contain only letters.")
            return False

        # Validate that Contact Number is a valid integer or "N/A"
        if guardian_contact != "N/A" and not guardian_contact.isdigit():
            messagebox.showerror("Invalid Input", "Error: Guardian's Contact Number should be a valid integer or 'N/A'.")
            return False

    # ==================== EDUCATIONAL BACKGROUND PAGE ====================
    # Validate Educational Information
    # Elementary
    Elem_school_name = Schl_Elem_Entry.get().strip()
    Elem_Address = Address_Elem_Entry.get().strip()
    Elem_year_grad = YR_Elem_Entry.get().strip()
    if Elem_school_name == "" or Elem_Address == "" or Elem_year_grad == "":
        messagebox.showerror("Missing Requirement", "Error: Elementary school information is incomplete.")
        return False
    if not Elem_year_grad.isdigit():
        messagebox.showerror("Invalid Input", "Error: Elementary Year Graduated must be an integer.")
        return False

    # Junior High
    jun_school_name = Schl_Junior_Entry.get().strip()
    jun_Address = Address_Junior_Entry.get().strip()
    jun_year_grad = YR_Junior_Entry.get().strip()
    if jun_school_name == "" or jun_Address == "" or jun_year_grad == "":
        messagebox.showerror("Missing Requirement", "Error: Junior High school information is incomplete.")
        return False
    if not jun_year_grad.isdigit():
        messagebox.showerror("Invalid Input", "Error: Junior High Year Graduated must be an integer.")
        return False

    # Senior High
    sen_school_name = Schl_Senior_Entry.get().strip()
    sen_Address = Address_Senior_Entry.get().strip()
    sen_strand = strand_Senior_Entry.get().strip()
    sen_year_grad = YR_Senior_Entry.get().strip()
    if sen_school_name == "" or  sen_Address == "" or sen_strand == "" or sen_year_grad == "":
        messagebox.showerror("Missing Requirement", "Error: Senior High school information is incomplete.")
        return False
    if not sen_year_grad.isdigit():
        messagebox.showerror("Invalid Input", "Error: Senior High Year Graduated must be an integer.")
        return False

    # College
    if if_transferee_var.get() == 0:
        col_school_name = Schl_College_Entry.get().strip()
        col_Address = Address_College_Entry.get().strip()
        col_year_grad = YR_College_Entry.get().strip()

        if col_school_name == "" or col_Address == "" or col_year_grad == "":
            messagebox.showerror("Missing Requirement", "Error: College information is incomplete.")
            return False
        if col_year_grad != "N/A" and not col_year_grad.isdigit():
            messagebox.showerror("Invalid Input", "Error: College Year Graduated must be an integer or 'N/A'.")
            return False
    
    return True

# Guardian Checkbox Toggler
def toggle_guardian_fields():
    state = "disabled" if same_chk_var.get() == 1 else "normal"
    if same_chk_var.get() == 1:
        Name_Guardian_Entry.delete(0, END)
        Address_Guardian_Entry.delete(0, END)
        Phone_Guardian_Entry.delete(0, END)
        Occupation_Guardian_Entry.delete(0, END)
        Rel_Guardian_Entry.delete(0, END)

    else:
        Name_Guardian_Entry.configure(state="normal")
        Address_Guardian_Entry.configure(state="normal")
        Phone_Guardian_Entry.configure(state="normal")
        Occupation_Guardian_Entry.configure(state="normal")
        Rel_Guardian_Entry.configure(state="normal")

        Name_Guardian_Entry.insert(0, "Enter Name")
        Address_Guardian_Entry.insert(0, "Enter Address")
        Phone_Guardian_Entry.insert(0, "Enter Contact Number")
        Occupation_Guardian_Entry.insert(0, "Enter Occupation")
        Rel_Guardian_Entry.insert(0, "Relationship")

    Name_Guardian_Entry.configure(state=state)
    Address_Guardian_Entry.configure(state=state)
    Phone_Guardian_Entry.configure(state=state)
    Occupation_Guardian_Entry.configure(state=state)
    Rel_Guardian_Entry.configure(state=state)

# Guardian Checkbox Toggler
def toggle_transferee_fields():
    state = "disabled" if if_transferee_var.get() == 1 else "normal"
    if if_transferee_var.get() == 1:
        Schl_College_Entry.delete(0, END)
        Address_College_Entry.delete(0, END)
        YR_College_Entry.delete(0, END)

    else:

        Schl_College_Entry.configure(state="normal")
        Address_College_Entry.configure(state="normal")
        YR_College_Entry.configure(state="normal")

    Schl_College_Entry.configure(state=state)
    Address_College_Entry.configure(state=state)
    YR_College_Entry.configure(state=state)

# Transferee Checkbox Toggler
def toggle_guardian_fields():
    state = "disabled" if same_chk_var.get() == 1 else "normal"
    if same_chk_var.get() == 1:
        Name_Guardian_Entry.delete(0, END)
        Address_Guardian_Entry.delete(0, END)
        Phone_Guardian_Entry.delete(0, END)
        Occupation_Guardian_Entry.delete(0, END)
        Rel_Guardian_Entry.delete(0, END)

    else:
        Name_Guardian_Entry.configure(state="normal")
        Address_Guardian_Entry.configure(state="normal")
        Phone_Guardian_Entry.configure(state="normal")
        Occupation_Guardian_Entry.configure(state="normal")
        Rel_Guardian_Entry.configure(state="normal")

    Name_Guardian_Entry.configure(state=state)
    Address_Guardian_Entry.configure(state=state)
    Phone_Guardian_Entry.configure(state=state)
    Occupation_Guardian_Entry.configure(state=state)
    Rel_Guardian_Entry.configure(state=state)

# ==================== Events ====================
def change_light_dark_mode_event(new_appearance_mode: str):
    ctk.set_appearance_mode(new_appearance_mode)

# ==================== Switch Window ====================
# Main Portal Window Switch Function
def GO_TO_PORTAL_WINDOW():
    window.destroy()
    subprocess.call(["python", "1_Portal_CTK.py"])

# Attendance Top Level Window Switch Function
def GO_TO_ATTENDANCE_WINDOW():
    subprocess.call(["python", "3_Attendance_Sheet_CTK.py"])

# //////////////////////////////////////////////////////////

# ==================== Navigation Buttons ====================
nav_frame = CTkFrame(window, fg_color="transparent")
nav_frame.pack(side="top", fill="x")

def show_frame(frame):
    frame.tkraise()

# Navigation Buttons
personal_btn = CTkButton(nav_frame, text="Personal Details", font=("Arial", 15, "bold"), command=lambda: show_frame(personal_frame))
personal_btn.pack(side="left", padx=5, pady=5)

family_btn = CTkButton(nav_frame, text="Family Background", font=("Arial", 15, "bold"), command=lambda: show_frame(family_frame))
family_btn.pack(side="left", padx=5, pady=5)

education_btn = CTkButton(nav_frame, text="Educational Background", font=("Arial", 15, "bold"), command=lambda: show_frame(education_frame))
education_btn.pack(side="left", padx=5, pady=5)

attendance_btn = CTkButton(nav_frame, text="Attendance", font=("Arial", 15, "bold"), command=GO_TO_ATTENDANCE_WINDOW)
attendance_btn.pack(side="left", padx=5, pady=5)

logout_btn = CTkButton(nav_frame, text="Logout", font=("Arial", 15, "bold"), command= GO_TO_PORTAL_WINDOW)
logout_btn.pack(side="right", padx=5, pady=5)

switch_light_button = CTkOptionMenu(nav_frame, values=["Dark", "Light"], command= change_light_dark_mode_event)
switch_light_button.pack(side="right", padx=5, pady=5)

# ==================== Main Container ====================
container = CTkFrame(window)
container.pack(fill="both", expand=True)

# ==================== Pages (Frames) ====================
personal_frame = CTkFrame(container)
family_frame = CTkFrame(container)
education_frame = CTkFrame(container)

for frame in (personal_frame, family_frame, education_frame):
    frame.place(x=0, y=0, relwidth=1, relheight=1)

# ==================== Footer ====================
present = datetime.now().strftime("%B %d, %Y - %I:%M %p")

footer = CTkLabel(window, text=f"Enrollment Form | Logged in as: Student | © {present}", font=("Arial", 12))
footer.pack(side="bottom", fill="x", pady=5)

# ==================== Buttons ====================
sub_frame = CTkFrame(window, border_width=2)
sub_frame.pack(side="bottom", fill="x")

submit_btn = CTkButton(sub_frame, text="Submit", font=("Arial", 15, "bold"), command=create_new_sheet)
submit_btn.pack(side="right", padx=5, pady=5)

load_btn = CTkButton(sub_frame, text="Load User Data", font=("Arial", 15, "bold"), command=load_user_data)
load_btn.pack(side="left", padx=5, pady=5)

edit_btn = CTkButton(sub_frame, text="Edit User Data", font=("Arial", 15, "bold"), command=edit_user_data)
edit_btn.pack(side="left", padx=5, pady=5)

update_btn = CTkButton(sub_frame, text="Update User Data", font=("Arial", 15, "bold"), command=update_user_data)
update_btn.pack(side="left", padx=5, pady=5)

# ==================== PERSONAL DETAILS PAGE ====================
# Scrollable Frame
scrollable_personal_frame = CTkScrollableFrame(personal_frame)
scrollable_personal_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
personal_header = CTkFrame(scrollable_personal_frame)
personal_header.pack(pady=20)
CTkLabel(personal_header, text="Lucena City", font=("Arial", 22)).pack()
CTkLabel(personal_header, text="STUDENT ENROLLMENT FORM", font=("Arial", 28, "bold")).pack()
CTkLabel(personal_header, text="Second Semester 2024-2025", font=("Arial", 22)).pack()

# Image and ID Section
top_section = CTkFrame(scrollable_personal_frame, fg_color= "transparent")
top_section.pack(padx=20, pady=10, fill="x")

left_top = CTkFrame(top_section, fg_color= "transparent")
left_top.pack(side="left", fill="both", expand=True)

Student_ID = CTkLabel(left_top, text="Student ID:", font=("Arial", 14))
Student_ID.grid(row=0, column=0, padx=40, pady=5, sticky="w")
Student_Entry = CTkEntry(left_top, width=250, font=("Arial", 14), placeholder_text="Enter Student ID", height= 35, fg_color= "transparent", bg_color= "transparent")
Student_Entry.grid(row=0, column=1, pady=5)

generate_student_id = CTkButton(left_top, text="Generate Student ID", font=("Arial", 14), command=generate_user_id)
generate_student_id.grid(row=0, column=2, padx=10, pady=5)

Course_Section = CTkLabel(left_top, text="Course/Section:", font=("Arial", 14))
Course_Section.grid(row=1, column=0, padx=40, pady=5, sticky="w")
sec_var = ["Select Section", "1A", "1B", "1C"]
Course_Section_Entry= CTkComboBox(left_top, width=250, font=("Arial", 14), height= 35, bg_color= "transparent", values=sec_var, state="readonly")
Course_Section_Entry.set("Select Section")
Course_Section_Entry.grid(row=1, column=1, pady=5)

lrn = CTkLabel(left_top, text="LRN:", font=("Arial", 14))
lrn.grid(row=2, column=0, padx=40, pady=5, sticky="w")
LRN_Entry= CTkEntry(left_top, width=250, font=("Arial", 14), placeholder_text="Enter LRN", height= 35, fg_color= "transparent", bg_color= "transparent")
LRN_Entry.grid(row=2, column=1, pady=5)

generate_student_lrn = CTkButton(left_top, text="Generate LRN", font=("Arial", 14), command=generate_user_lrn)
generate_student_lrn.grid(row=2, column=2, padx=10, pady=5)

right_top = CTkFrame(top_section, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
right_top.pack(side="right", padx=10)

# Upload Picture 2x2
def upload_image():
    """Uploads an image from the user's computer to the window."""
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.jpeg *.png")])
    if file_path:
        with Image.open(file_path) as img:
            img = img.resize((200, 200), Image.Resampling.LANCZOS)
            img_ctk = CTkImage(light_image=img, dark_image=img, size=(200, 200))
            picture_label.configure(image=img_ctk, text="")
            picture_label.image = img_ctk

picture_frame = CTkFrame(right_top, width=200, height=200, border_color="black", border_width=6, corner_radius=10, fg_color="transparent")
picture_frame.pack()
picture_frame.pack_propagate(False)  # Prevent frame from resizing

# Clickable label to display/upload image
picture_label = CTkLabel(
    picture_frame,
    text="Upload\n2x2 Picture",
    justify="center",
    font=("Arial", 12),
    bg_color="white",
    text_color="black"
)
picture_label.pack(fill="both", expand=True)
picture_label.bind("<Button-1>", lambda e: upload_image())

# Show personal details first
show_frame(personal_frame)

# Personal Information Title
personal_title_info =CTkLabel(scrollable_personal_frame, text= "PERSONAL INFORMATION", font= ("Id Inter", 25))
personal_title_info.pack(padx=20, pady=10, anchor=W)

# Personal Details Section
personal_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
personal_details.pack(padx=20, pady=10, fill="x")

# First Row
first_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
first_row_personal_details_frame.grid(row=0, column=0, columnspan=3, sticky="w", padx=20, pady=7)

surname_label = CTkLabel(first_row_personal_details_frame, text="Surname:", font=("Arial", 14), bg_color="transparent")
surname_label.grid(row=0, column=0, padx=20, pady=7, sticky="w")
Surname_Entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Surname", height= 35, fg_color= "transparent", bg_color= "transparent")
Surname_Entry.grid(row=1, column=0, padx=20, sticky="wn")

firstname_label = CTkLabel(first_row_personal_details_frame, text="Firstname:", font=("Arial", 14), bg_color="transparent")
firstname_label.grid(row=0, column=1, padx=20, pady=7, sticky="w")
FirstName_Entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Firstname", height= 35, fg_color= "transparent", bg_color= "transparent")
FirstName_Entry.grid(row=1, column=1, padx=20, sticky="wn")

middle_label = CTkLabel(first_row_personal_details_frame, text="Middle Initial:", font=("Arial", 14), bg_color="transparent")
middle_label.grid(row=0, column=2, padx=20, pady=7, sticky="w")
Middle_Entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Middle Initial", height= 35, fg_color= "transparent", bg_color= "transparent")
Middle_Entry.grid(row=1, column=2, padx=20, sticky="wn")

# Second Row
second_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
second_row_personal_details_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=20, pady=7)

gender_label = CTkLabel(second_row_personal_details_frame, text="Gender:", font=("Arial", 14), bg_color="transparent")
gender_label.grid(row=1, column=0, sticky="w", padx=40, pady=7)
gender_frame = CTkFrame(second_row_personal_details_frame, bg_color="transparent", fg_color= "transparent")
gender_frame.grid(row=2, column=0, sticky="w", padx=40)

gender_var = IntVar(value=0)

male_checkbox = CTkRadioButton(gender_frame, text="Male", variable=gender_var, value=1)
male_checkbox.grid(row=0, column=0)
female_checkbox = CTkRadioButton(gender_frame, text="Female", variable=gender_var, value=2)
female_checkbox.grid(row=0, column=1)
none_binary_checkbox = CTkRadioButton(gender_frame, text="Prefer not to answer", variable=gender_var, value=3)
none_binary_checkbox.grid(row=0, column=2)

age_label = CTkLabel(second_row_personal_details_frame, text="Age:", font=("Arial", 14), bg_color="transparent")
age_label.grid(row=1, column=1, sticky="w", padx=20, pady=7)
age_values = ["Select Age"] + [str(i) for i in range(1, 101)]
age_box = CTkComboBox(second_row_personal_details_frame, values=age_values, width=150, height= 35, bg_color= "transparent", state="normal")
age_box.set("Select Age")
age_box.grid(row=2, column=1, sticky="wn", padx=20)

birthdate_label = CTkLabel(second_row_personal_details_frame, text="Birthdate (MM/DD/YYYY):", font=("Arial", 14), bg_color="transparent")
birthdate_label.grid(row=1, column=2, sticky="w", padx=20, pady=7)
birthdate_frame = CTkFrame(second_row_personal_details_frame, bg_color="transparent", fg_color= "transparent")
birthdate_frame.grid(row=2, column=2, sticky="w", padx=20)

months = [
          "MM", "January", "February", "March", "April", 
          "May", "June", "July", "August", "September", 
          "October", "November", "December"
          ]
month_box = CTkComboBox(birthdate_frame, values=months, width=120, height= 35, bg_color= "transparent", state="normal")
month_box.set("MM")
month_box.grid(row=0, column=0, padx=3)

days = ["DD"] + [str(x) for x in range(1, 32)]
day_box = CTkComboBox(birthdate_frame, values=days, width=80, height= 35, bg_color= "transparent", state="normal")
day_box.set("DD")
day_box.grid(row=0, column=1, padx=3)

years = ["YYYY"] + [str(x) for x in range(1900, datetime.now().year + 11)]
year_box = CTkComboBox(birthdate_frame, values=years, width=150, height= 35, bg_color= "transparent", state="normal")
year_box.set("YYYY")
year_box.grid(row=0, column=2, padx=3)

# Third Row
third_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
third_row_personal_details_frame.grid(row=3, column=0, columnspan=3, sticky="we", padx=20, pady=7)

birthplace_label = CTkLabel(third_row_personal_details_frame, text="Birthplace City:", font=("Arial", 14), bg_color="transparent")
birthplace_label.grid(row=0, column=0, sticky="w", padx=20, pady=7)
Birthplace_Entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Birthplace", height= 35, fg_color= "transparent", bg_color= "transparent")
Birthplace_Entry.grid(row=1, column=0, padx=20, sticky="wn")

nationality_label = CTkLabel(third_row_personal_details_frame, text="Nationality:", font=("Arial", 14), bg_color="transparent")
nationality_label.grid(row=0, column=1, sticky="w", padx=20, pady=7)
Nationality_Entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Nationality", height= 35, fg_color= "transparent", bg_color= "transparent")
Nationality_Entry.grid(row=1, column=1, padx=20, sticky="wn")

religion_label = CTkLabel(third_row_personal_details_frame, text="Religion:", font=("Arial", 14), bg_color="transparent")
religion_label.grid(row=0, column=2, sticky="w", padx=20, pady=7)
religions = [
             "Select Religion", "Roman Catholic", 
             "Iglesia ni Cristo", "Born Again", 
             "Protestant", "Muslim", "Other"
             ]
religion_box = CTkComboBox(third_row_personal_details_frame, values=religions, width=150, height= 35, bg_color= "transparent", state="readonly")
religion_box.set("Select Religion")
religion_box.grid(row=1, column=2, padx=20)

marital_status_label = CTkLabel(third_row_personal_details_frame, text="Marital Status:", font=("Arial", 14), bg_color="transparent")
marital_status_label.grid(row=0, column=3, sticky="w", padx=20, pady=7)
statuses = [
             "Select Marital Status", "Single", 
             "Married", "Widowed", "Separated"
            ]
marital_status_box = CTkComboBox(third_row_personal_details_frame, values=statuses, width=180, height= 35, bg_color= "transparent", state="readonly")
marital_status_box.set("Select Marital Status")
marital_status_box.grid(row=1, column=3, padx=20)

language_label = CTkLabel(third_row_personal_details_frame, text="Language Spoken:", font=("Arial", 14), bg_color="transparent")
language_label.grid(row=0, column=4, sticky="w", padx=20, pady=7)
Language_Entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Language", height= 35, fg_color= "transparent", bg_color= "transparent")
Language_Entry.grid(row=1, column=4, padx=20, sticky="wn")

empty_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=4, column=0, columnspan=3, sticky="we", padx=20, pady=7)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

# Address Title
Address_title_info =CTkLabel(scrollable_personal_frame, text= "Address", font= ("Id Inter", 25))
Address_title_info.pack(padx=20, pady=10, anchor=W)

# Address Details Section
Address_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
Address_details.pack(padx=20, pady=10, fill="x")

first_row_Address_details_frame = CTkFrame(Address_details, bg_color="transparent", fg_color= "transparent")
first_row_Address_details_frame.grid(row=0, column=0, columnspan=3, sticky="we", padx=20, pady=7)

street_label = CTkLabel(first_row_Address_details_frame, text="Street:", font=("Arial", 14), bg_color="transparent")
street_label.grid(row=0, column=0, padx=20, pady=7, sticky="w")
Street_Entry = CTkEntry(first_row_Address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Street", height= 35, fg_color= "transparent", bg_color= "transparent")
Street_Entry.grid(row=1, column=0, padx=20, sticky="wn")

brgy_label = CTkLabel(first_row_Address_details_frame, text="Barangay:", font=("Arial", 14), bg_color="transparent")
brgy_label.grid(row=0, column=1, padx=20, pady=7, sticky="w")
BRGY_Entry = CTkEntry(first_row_Address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Barangay", height= 35, fg_color= "transparent", bg_color= "transparent")
BRGY_Entry.grid(row=1, column=1, padx=20, sticky="wn")

city_label = CTkLabel(first_row_Address_details_frame, text="City/Municipality:", font=("Arial", 14), bg_color="transparent")
city_label.grid(row=0, column=2, padx=20, pady=7, sticky="w")
City_Entry = CTkEntry(first_row_Address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter City", height= 35, fg_color= "transparent", bg_color= "transparent")
City_Entry.grid(row=1, column=2, padx=20, sticky="wn")

zip_code_label = CTkLabel(first_row_Address_details_frame, text="Zip Code:", font=("Arial", 14), bg_color="transparent")
zip_code_label.grid(row=0, column=3, padx=20, pady=7, sticky="w")
ZipCode_Entry = CTkEntry(first_row_Address_details_frame, width=170, font=("Arial", 14), placeholder_text="Enter Zip Code", height= 35, fg_color= "transparent", bg_color= "transparent")
ZipCode_Entry.grid(row=1, column=3, padx=20, sticky="wn")

province_label = CTkLabel(first_row_Address_details_frame, text="Province:", font=("Arial", 14), bg_color="transparent")
province_label.grid(row=2, column=0, padx=20, pady=7, sticky="w")
Province_Entry = CTkEntry(first_row_Address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Province", height= 35, fg_color= "transparent", bg_color= "transparent")
Province_Entry.grid(row=3, column=0, padx=20, sticky="wn")

country_label = CTkLabel(first_row_Address_details_frame, text="Country:", font=("Arial", 14), bg_color="transparent")
country_label.grid(row=2, column=1, padx=20, pady=7, sticky="w")
Country_Entry = CTkEntry(first_row_Address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Country", height= 35, fg_color= "transparent", bg_color= "transparent")
Country_Entry.grid(row=3, column=1, padx=20, sticky="wn")

empty_frame = CTkFrame(Address_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=1, column=0, columnspan=3, sticky="we", padx=20, pady=6)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

# Contact Information Title
contact_title_info =CTkLabel(scrollable_personal_frame, text= "CONTACT INFORMATION", font= ("Id Inter", 25))
contact_title_info.pack(padx=20, pady=10, anchor=W)

# Contact Information Details Section
contact_info_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
contact_info_details.pack(padx=20, pady=10, fill="x")

first_row_contact_info_details_frame = CTkFrame(contact_info_details, bg_color="transparent", fg_color= "transparent")
first_row_contact_info_details_frame.grid(row=0, column=0, columnspan=3, sticky="we", padx=35, pady=7)

email_label = CTkLabel(first_row_contact_info_details_frame, text="Email:", font=("Arial", 14), bg_color="transparent")
email_label.grid(row=0, column=0, padx=10, pady=7, sticky="w")
Email_Entry = CTkEntry(first_row_contact_info_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Email", height= 35, fg_color= "transparent", bg_color= "transparent")
Email_Entry.grid(row=0, column=1, padx=10, pady=7, sticky="wn")
Email_Entry.bind("<FocusOut>", lambda event: validating_user_email())

num_label = CTkLabel(first_row_contact_info_details_frame, text="Contact Number:", font=("Arial", 14), bg_color="transparent")
num_label.grid(row=0, column=2, padx=10, pady=7, sticky="w")
Num_Entry = CTkEntry(first_row_contact_info_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
Num_Entry.grid(row=0, column=3, padx=10, pady=7, sticky="wn")

empty_frame = CTkFrame(contact_info_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=1, column=0, columnspan=3, sticky="we", padx=20, pady=7)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)


# ==================== FAMILY BACKGROUND PAGE ====================
# Scrollable Frame
scrollable_family_background_frame = CTkScrollableFrame(family_frame)
scrollable_family_background_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
family_background_header = CTkFrame(scrollable_family_background_frame)
family_background_header.pack(pady=20)
CTkLabel(family_background_header, text="FAMILY BACKGROUND", font=("Arial", 28, "bold")).pack()

# Parent's Information Title
family_background_title_info =CTkLabel(scrollable_family_background_frame, text= "PARENT'S INFORMATION", font= ("Id Inter", 25))
family_background_title_info.pack(padx=20, pady=10, anchor=W)

# Parent's Details Section
parents_details = CTkFrame(scrollable_family_background_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
parents_details.pack(padx=20, pady=10, fill="x")

row_column_parents_details_frame = CTkFrame(parents_details, bg_color="transparent", fg_color= "transparent")
row_column_parents_details_frame.pack(anchor = CENTER, pady=30)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

father_label = CTkLabel(row_column_parents_details_frame, text="FATHER", font=("Arial", 14, "bold"), bg_color="transparent")
father_label.grid(row=0, column=1, padx=10, pady=7)

mother_label = CTkLabel(row_column_parents_details_frame, text="MOTHER", font=("Arial", 14, "bold"), bg_color="transparent")
mother_label.grid(row=0, column=2, padx=10, pady=7)

name_label = CTkLabel(row_column_parents_details_frame, text="Name:", font=("Arial", 14), bg_color="transparent")
name_label.grid(row=1, column=0, padx=10, pady=7, sticky="w")

Name_Father_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
Name_Father_Entry.grid(row=1, column=1, padx=10, pady=7)

Name_Mother_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
Name_Mother_Entry.grid(row=1, column=2, padx=10, pady=7)

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=2, column=0, padx=10, pady=7, sticky="w")

Address_Father_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Father_Entry.grid(row=2, column=1, padx=10, pady=7)

Address_Mother_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Mother_Entry.grid(row=2, column=2, padx=10, pady=7)

occupation_label = CTkLabel(row_column_parents_details_frame, text="Occupation:", font=("Arial", 14), bg_color="transparent")
occupation_label.grid(row=3, column=0, padx=10, pady=7, sticky="w")

Occupation_Father_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
Occupation_Father_Entry.grid(row=3, column=1, padx=10, pady=7)

Occupation_Mother_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
Occupation_Mother_Entry.grid(row=3, column=2, padx=10, pady=7)

phon_num_label = CTkLabel(row_column_parents_details_frame, text="Contact Number:", font=("Arial", 14), bg_color="transparent")
phon_num_label.grid(row=4, column=0, padx=10, pady=7, sticky="w")

Phone_Father_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
Phone_Father_Entry.grid(row=4, column=1, padx=10, pady=7)

Phone_Mother_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
Phone_Mother_Entry.grid(row=4, column=2, padx=10, pady=7)

guardian_label = CTkLabel(row_column_parents_details_frame, text="GUARDIAN", font=("Arial", 14, "bold"), bg_color="transparent")
guardian_label.grid(row=5, column=1, padx=10, pady=7)

same_chk_var = IntVar()
same_chk_box = CTkCheckBox(row_column_parents_details_frame, text="Same as Father and/or Mother", variable=same_chk_var)
same_chk_box.grid(row=5, column=2, padx=10, pady=7, sticky="w")
same_chk_box.configure(command=toggle_guardian_fields)

name_label = CTkLabel(row_column_parents_details_frame, text="Name:", font=("Arial", 14), bg_color="transparent")
name_label.grid(row=6, column=0, padx=10, pady=7, sticky="w")

Name_Guardian_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
Name_Guardian_Entry.grid(row=6, column=1, padx=10, pady=7)

rel_label = CTkLabel(row_column_parents_details_frame, text="Relationship to Student:", font=("Arial", 14), bg_color="transparent")
rel_label.grid(row=6, column=2, padx=10, pady=7, sticky="s")

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=7, column=0, padx=10, pady=7, sticky="w")

Address_Guardian_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Guardian_Entry.grid(row=7, column=1, padx=10, pady=7)

Rel_Guardian_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Relationship", height= 35, fg_color= "transparent", bg_color= "transparent")
Rel_Guardian_Entry.grid(row=7, column=2, padx=10, pady=7)

occupation_label = CTkLabel(row_column_parents_details_frame, text="Occupation:", font=("Arial", 14), bg_color="transparent")
occupation_label.grid(row=8, column=0, padx=10, pady=7, sticky="w")

Occupation_Guardian_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
Occupation_Guardian_Entry.grid(row=8, column=1, padx=10, pady=7)

phon_num_label = CTkLabel(row_column_parents_details_frame, text="Contact Number:", font=("Arial", 14), bg_color="transparent")
phon_num_label.grid(row=9, column=0, padx=10, pady=7, sticky="w")

Phone_Guardian_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
Phone_Guardian_Entry.grid(row=9, column=1, padx=10, pady=7)

toggle_guardian_fields()

# ==================== EDUCATIONAL BACKGROUND PAGE ====================
# Scrollable Frame
scrollable_educ_background_frame = CTkScrollableFrame(education_frame)
scrollable_educ_background_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
educ_background_header = CTkFrame(scrollable_educ_background_frame)
educ_background_header.pack(pady=20)
CTkLabel(educ_background_header, text="EDUCATIONAL BACKGROUND", font=("Arial", 28, "bold")).pack()

# Educ's Information Title
family_background_title_info =CTkLabel(scrollable_educ_background_frame, text= "EDUCATIONAL INFORMATION", font= ("Id Inter", 25))
family_background_title_info.pack(padx=20, pady=10, anchor=W)

# Educ's Details Section
educ_details = CTkFrame(scrollable_educ_background_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
educ_details.pack(padx=20, pady=10, fill="x")

row_column_parents_details_frame = CTkFrame(educ_details, bg_color="transparent", fg_color= "transparent")
row_column_parents_details_frame.pack(anchor = CENTER, pady=30)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

Elem_label = CTkLabel(row_column_parents_details_frame, text="ElemENTARY", font=("Arial", 14, "bold"), bg_color="transparent")
Elem_label.grid(row=0, column=1, padx=10, pady=7)

Schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
Schl_label.grid(row=1, column=0, padx=10, pady=7, sticky="w")
Schl_Elem_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
Schl_Elem_Entry.grid(row=1, column=1, padx=10, pady=7)

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=2, column=0, padx=10, pady=7, sticky="w")
Address_Elem_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Elem_Entry.grid(row=2, column=1, padx=10, pady=7)

YR_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
YR_label.grid(row=3, column=0, padx=10, pady=7, sticky="w")
YR_Elem_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
YR_Elem_Entry.grid(row=3, column=1, padx=10, pady=7)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=2)

Junior_label = CTkLabel(row_column_parents_details_frame, text="JUNIOR HIGH SCHOOL", font=("Arial", 14, "bold"), bg_color="transparent")
Junior_label.grid(row=0, column=3, padx=10, pady=7)

Schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
Schl_label.grid(row=1, column=2, padx=10, pady=7, sticky="w")
Schl_Junior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
Schl_Junior_Entry.grid(row=1, column=3, padx=10, pady=7)

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=2, column=2, padx=10, pady=7, sticky="w")
Address_Junior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Junior_Entry.grid(row=2, column=3, padx=10, pady=7)

YR_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
YR_label.grid(row=3, column=2, padx=10, pady=7, sticky="w")
YR_Junior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
YR_Junior_Entry.grid(row=3, column=3, padx=10, pady=7)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=4, column=0)

Senior_label = CTkLabel(row_column_parents_details_frame, text="SENIOR HIGH SCHOOL", font=("Arial", 14, "bold"), bg_color="transparent")
Senior_label.grid(row=4, column=1, padx=10, pady=7)

Schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
Schl_label.grid(row=5, column=0, padx=10, pady=7, sticky="w")
Schl_Senior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
Schl_Senior_Entry.grid(row=5, column=1, padx=10, pady=7)

strand_label = CTkLabel(row_column_parents_details_frame, text="Strand:", font=("Arial", 14), bg_color="transparent")
strand_label.grid(row=6, column=0, padx=10, pady=7, sticky="w")
strand_Senior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Strand", height= 35, fg_color= "transparent", bg_color= "transparent")
strand_Senior_Entry.grid(row=6, column=1, padx=10, pady=7)

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=7, column=0, padx=10, pady=7, sticky="w")
Address_Senior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_Senior_Entry.grid(row=7, column=1, padx=10, pady=7)

YR_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
YR_label.grid(row=8, column=0, padx=10, pady=7, sticky="wn")
YR_Senior_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
YR_Senior_Entry.grid(row=8, column=1, padx=10, pady=7, sticky="n")

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=4, column=2)

College_label = CTkLabel(row_column_parents_details_frame, text="COLLEGE", font=("Arial", 14, "bold"), bg_color="transparent")
College_label.grid(row=4, column=3, padx=10, pady=7, sticky="w")

if_transferee_var = IntVar()
if_transferee_chk_box = CTkCheckBox(educ_details, text="Check if you are not a transferee", variable=if_transferee_var)
if_transferee_chk_box.place(x=834, y=230)
if_transferee_chk_box.configure(command=toggle_transferee_fields)

Schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
Schl_label.grid(row=5, column=2, padx=10, pady=7, sticky="w")
Schl_College_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
Schl_College_Entry.grid(row=5, column=3, padx=10, pady=7)

Address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
Address_label.grid(row=6, column=2, padx=10, pady=7, sticky="w")
Address_College_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
Address_College_Entry.grid(row=6, column=3, padx=10, pady=7)

YR_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
YR_label.grid(row=7, column=2, padx=10, pady=7, sticky="w")
YR_College_Entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
YR_College_Entry.grid(row=7, column=3, padx=10, pady=7)

# ==================== Window Starter ====================
window.mainloop()