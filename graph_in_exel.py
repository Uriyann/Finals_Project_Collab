import tkinter as tk
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import re
from tkinter import Image, PhotoImage
import os
from PIL import ImageTk, Image


window = tk.Tk()
window.title("score tracker")
window.geometry("600x300")
window.configure(bg= "gray5")
window.resizable(False, False)

# ✅ Step 1: Create the original Excel file with grades

def create_excel_file():

    wb = Workbook()
    ws = wb.active
    ws.title = "grades"
    ws.append(["Subject", "Grade"])
    wb.save("grades.xlsx")
    print("✅ grades.xlsx created!")

def create_grades_excel():

    subject = subject_entry.get()
    score= int(score_entry.get())

    wb = load_workbook("grades.xlsx")
    if "grades" not in wb.sheetnames:
        create_excel_file()
    else:
        ws = wb["grades"]
        ws.append([subject, score])
    

    
  # Optional formatting
    messagebox.showinfo(title="Success",message= "Data saved successfully!")

    subject_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)

    subject_entry.insert(0, "subject")
    score_entry.insert(0, "Score")


    wb.save("grades.xlsx")
    print("✅ grades.xlsx created!")

# ✅ Step 2: Read data from Excel and make chart

def show_chart_image():
    chart_path = "grades_chart.png"
    if os.path.exists(chart_path):
        # Open new window
        img_window = Toplevel()
        img_window.title("Grades Chart")
        
        # Load and display image
        img = Image.open(chart_path)
        img = img.resize((500, 300))  # Resize if needed
        photo = ImageTk.PhotoImage(img)

        label = Label(img_window, image=photo)
        label.image = photo  # Keep reference!
        label.pack()
    else:
        messagebox.showerror("Error", "Chart image not found. Please generate it first.")

def create_bar_chart_from_excel():
    try:
        # Load the workbook and worksheet
        wb = load_workbook("grades.xlsx")
        ws = wb.active

        # Initialize lists to store subjects and grades
        subjects = []
        grades = []

        # Read data from the worksheet
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
            if row[0] is not None and row[1] is not None:  # Ensure both subject and grade are not None
                subjects.append(row[0])
                grades.append(row[1])

        # Check if there is data to plot
        if not subjects or not grades:
            messagebox.showerror("Error", "No data available to create the chart.")
            return

        # Create the bar chart
        plt.bar(subjects, grades, color='skyblue')
        plt.title('Student Grades')
        plt.ylabel('Scores')
        plt.tight_layout()

        # Save the chart as an image
        plt.savefig("grades_chart.png")
        plt.close()
        print("✅ Chart image saved as grades_chart.png")
        messagebox.showinfo("Success", "Chart created and saved as 'grades_chart.png'.")

    except FileNotFoundError:
        messagebox.showerror("Error", "The file 'grades.xlsx' does not exist. Please create it first.")
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")



# ✅ Step 3: Insert chart image into new Excel file
def insert_chart_to_excel():
    try:
        # Ensure the chart image exists

        chart_path = "grades_chart.png"
        plt.savefig(chart_path)
        plt.close()

        chart_path = "grades_chart.png"
        plt.savefig("grades_chart.png")
        if not os.path.exists(chart_path):
            raise FileNotFoundError(f"Chart image '{chart_path}' not found. Please create the chart first.")

        # Create a new workbook and worksheet
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Grades with Chart"

        # Add a title to the worksheet
        new_ws['A1'] = "Student Grades"

        # Insert the chart image
        img = Image(chart_path)
        new_ws.add_image(img, "A3")

        # Save the new workbook
        new_wb.save("grades_with_chart.xlsx")
        print("✅ New Excel file saved as grades_with_chart.xlsx")
        messagebox.showinfo("Success", "Chart inserted into Excel file successfully!")

    except FileNotFoundError as e:
        messagebox.showerror("Error", str(e))
    except Exception as e:
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")

# 🟩 Run all 

mainframe= tk.Frame(window, bg= "gray25")
mainframe.place( anchor= "center", relx= 0.5, rely= 0.5, relwidth= 0.8, relheight= 0.8)


user_frame= tk.LabelFrame(mainframe,bg= "gray30", )

user_frame.pack(pady=20, padx=20, fill= "both", expand= True)

title_frame= tk.Frame(user_frame, bg= "gray30")
title_frame.pack()

user= tk.Label(title_frame, text= "User Information", font= ("Arial", 20, "bold"),bg= "gray30", fg= "white")


user.pack()

input_frame= tk.Frame(user_frame, bg= "gray30")
input_frame.pack(fill= "both", expand= True)

subject_label= tk.Label(input_frame, text= "Subject:",font= ("Arial", 15, "bold"), fg= "white",bg= "gray30")
subject_label.grid(row= 0, column= 0, padx= 10, pady= 10, sticky= "w")

subject_entry= tk.Entry(input_frame, width= 30)
subject_entry.grid(row= 0, column= 1, padx= 10, pady= 10)

score= tk.Label(input_frame, text= "grade:", font= ("Arial", 15, "bold" ), fg= "white",bg = "gray30")
score.grid(row= 1, column= 0, padx= 10, pady= 10, sticky= "w")

score_entry= tk.Entry(input_frame, width= 30)
score_entry.grid(row= 1, column= 1, padx= 10, pady= 10)

enter_button= tk.Button(input_frame, text= "Enter", font= ("Arial", 15), bg= "yellow2",command= create_grades_excel)
enter_button.grid(row= 2, column= 0, padx= 10, pady= 10)

show_data_button= tk.Button(input_frame, text= "Show Data", font= ("Arial", 15), bg= "dodger blue", command= create_bar_chart_from_excel)   
show_data_button.grid(row= 2, column= 1, padx= 10, pady= 10)

show_graph_button= tk.Button(input_frame, text= "Show Graph", font= ("Arial", 15), bg= "dodger blue", command= show_chart_image)
show_graph_button.grid(row= 2, column= 2, padx= 10, pady= 10)   

create_excel_file()
window.mainloop()

