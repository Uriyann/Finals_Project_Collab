from customtkinter import *
import customtkinter as ctk
from datetime import datetime
from tkinter import filedialog
from PIL import Image
from tkinter import messagebox
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ==================== Functions ====================
# Creating a new sheet in the workbook and saving it
def create_new_sheet():
    try:

        sections = Course_Section_enrty.get()
        if sections not in ["1A", "1B", "1C"]:
            messagebox.showerror(title="Error", message="Please select a valid section.")
            return
        
        new_sheet_name = f"{sections} Enrollment Information "
        file_path_to_excel = "user_account_data.xlsx"

        wb = load_workbook(file_path_to_excel)
        if new_sheet_name in wb.sheetnames:
            new_sheet = wb[new_sheet_name]    

        else:
            new_sheet = wb.create_sheet(new_sheet_name)

        new_headers = [
                "Student ID", "Course/Section", "LRN", "", "Surname", "Firstname", 
                "Middle Initial", "","Gender", "Age", "Birthdate", "Birthplace", 
                "Nationality", "Religion", "Marital Status", "Language Spoken", "", 
                "Street", "Barangay", "City", "Zip Code", "Province", "Country", "",
                "Email", "Contact Number"
                    ]
        new_sheet.append(new_headers)

        student_name_data = [
            surname_entry.get(), firstname_entry.get(), middle_entry.get(), ""
                ]

        student_info_data = [
                Student_entry.get(), Course_Section_enrty.get(), lrn_entry.get(), ""
                ]

        student_personal_detail_data = [
                "Male" if male_var.get() else "Female" if female_var.get() else "Prefer not to answer", 
                age_box.get(), f"{month_box.get()} {day_box.get()}, {year_box.get()}",
                birthplace_entry.get(), nationality_entry.get(), religion_box.get(),
                marital_status_box.get(), language_entry.get(), "",
                street_entry.get(), brgy_entry.get(), city_entry.get(), zip_code_entry.get(),
                province_entry.get(), country_entry.get(), "",
                email_entry.get(), num_entry.get()
                ]
        
        new_sheet.append(student_name_data + student_info_data + student_personal_detail_data)
        format_excel()

        wb.save(file_path_to_excel)
        messagebox.showinfo(title= "Success", message= "Data saved successfully!")

    except FileNotFoundError:
        messagebox.showerror(title= "Error", message= "User data file not found.")

# Validating the User per Email
def validating_user_email():
    try:

        email = email_entry.get()
        if not email:
            messagebox.showerror(title="Error", message="Please enter an email address.")
            return

        wb = load_workbook("user_account_data.xlsx")
        account_data_sheet = wb["Userdata"]
        for row in account_data_sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == email:
                surname_entry.delete(0, END)
                surname_entry.insert(0, row[1])

                firstname_entry.delete(0, END)
                firstname_entry.insert(0, row[0])

                email_entry.delete(0, END)
                email_entry.insert(0, row[2])
                return 
        
        messagebox.showerror(title="Error", message="User not found.")
            
    except Exception as e:
        messagebox.showerror(title="Error", message=f"An error occurred: {e}")

# Format Fixer Function
def format_excel():
    sections = Course_Section_enrty.get()

    new_sheet_name = f"{sections} Enrollment Information "

    wb = load_workbook("user_account_data.xlsx")
    ws = wb[new_sheet_name]

    # Bold Headers
    for cells in ws[1]:
        cells.font = Font(bold=True)

    # Auto Column Width
    for cols in ws.columns:
        max_length = max(len(str(cells.value)) for cell in cols)
        col_letter = get_column_letter(cols[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("user_account_data.xlsx")

# ==================== Window Setup ====================
window = CTk()
window.title("Student Enrollment Form")
window.geometry('1300x825')
window.resizable(False, False)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ==================== Events ====================
def change_light_dark_mode_event(new_appearance_mode: str):
    ctk.set_appearance_mode(new_appearance_mode)

# ==================== Switch Window ====================
# Main Portal Window Switch Function
def GO_TO_PORTAL_WINDOW():
    window.destroy()
    subprocess.call(["python", "1_Portal_CTK.py"])

# //////////////////////////////////////////////////////////

# ==================== Navigation Buttons ====================
nav_frame = CTkFrame(window, fg_color="transparent")
nav_frame.pack(side="top", fill="x")

def show_frame(frame):
    frame.tkraise()

# Navigation Buttons
personal_btn = CTkButton(nav_frame, text="Personal Details", font=("Arial", 15, "bold"), command=lambda: show_frame(personal_frame))
personal_btn.pack(side="left", padx=5, pady=5)

family_btn = CTkButton(nav_frame, text="Family Background", font=("Arial", 15, "bold"), command=lambda: show_frame(family_frame))
family_btn.pack(side="left", padx=5, pady=5)

education_btn = CTkButton(nav_frame, text="Educational Background", font=("Arial", 15, "bold"), command=lambda: show_frame(education_frame))
education_btn.pack(side="left", padx=5, pady=5)

logout_btn = CTkButton(nav_frame, text="Logout", font=("Arial", 15, "bold"), command= GO_TO_PORTAL_WINDOW)
logout_btn.pack(side="right", padx=5, pady=5)

switch_light_button = CTkOptionMenu(nav_frame, values=["Dark", "Light"], command= change_light_dark_mode_event)
switch_light_button.pack(side="right", padx=5, pady=5)

# ==================== Main Container ====================
container = CTkFrame(window)
container.pack(fill="both", expand=True)

# ==================== Pages (Frames) ====================
personal_frame = CTkFrame(container)
family_frame = CTkFrame(container)
education_frame = CTkFrame(container)

for frame in (personal_frame, family_frame, education_frame):
    frame.place(x=0, y=0, relwidth=1, relheight=1)

# ==================== Buttons ====================
sub_frame = CTkFrame(window, fg_color="transparent")
sub_frame.pack(side="bottom", fill="x")

submit_btn = CTkButton(sub_frame, text="Submit", font=("Arial", 15, "bold"), command=create_new_sheet)
submit_btn.pack(side="right", padx=5, pady=5)

# ==================== PERSONAL DETAILS PAGE ====================
# Scrollable Frame
scrollable_personal_frame = CTkScrollableFrame(personal_frame)
scrollable_personal_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
personal_header = CTkFrame(scrollable_personal_frame)
personal_header.pack(pady=20)
CTkLabel(personal_header, text="Lucena City", font=("Arial", 22)).pack()
CTkLabel(personal_header, text="STUDENT ENROLLMENT FORM", font=("Arial", 28, "bold")).pack()
CTkLabel(personal_header, text="Second Semester 2024-2025", font=("Arial", 22)).pack()

# Image and ID Section
top_section = CTkFrame(scrollable_personal_frame, fg_color= "transparent")
top_section.pack(padx=20, pady=10, fill="x")

left_top = CTkFrame(top_section, fg_color= "transparent")
left_top.pack(side="left", fill="both", expand=True)

Student_ID = CTkLabel(left_top, text="Student ID:", font=("Arial", 14))
Student_ID.grid(row=0, column=0, padx=40, pady=5, sticky="w")
Student_entry = CTkEntry(left_top, width=250, font=("Arial", 14), placeholder_text="Enter Student ID", height= 35, fg_color= "transparent", bg_color= "transparent")
Student_entry.grid(row=0, column=1, pady=5)

Course_Section = CTkLabel(left_top, text="Course/Section:", font=("Arial", 14))
Course_Section.grid(row=1, column=0, padx=40, pady=5, sticky="w")
sec_var = ["Select Section", "1A", "1B", "1C"]
Course_Section_enrty= CTkComboBox(left_top, width=250, font=("Arial", 14), height= 35, bg_color= "transparent", values=sec_var, state="readonly")
Course_Section_enrty.set("Select Section")
Course_Section_enrty.grid(row=1, column=1, pady=5)

lrn = CTkLabel(left_top, text="LRN:", font=("Arial", 14))
lrn.grid(row=2, column=0, padx=40, pady=5, sticky="w")
lrn_entry= CTkEntry(left_top, width=250, font=("Arial", 14), placeholder_text="Enter LRN", height= 35, fg_color= "transparent", bg_color= "transparent")
lrn_entry.grid(row=2, column=1, pady=5)

right_top = CTkFrame(top_section, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
right_top.pack(side="right", padx=10)

# Upload Picture 2x2
def upload_image():
    """Uploads an image from the user's computer to the window."""
    file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.jpeg *.png")])
    if file_path:
        with Image.open(file_path) as img:
            img = img.resize((200, 200), Image.Resampling.LANCZOS)
            img_ctk = CTkImage(light_image=img, dark_image=img, size=(200, 200))
            picture_label.configure(image=img_ctk, text="")
            picture_label.image = img_ctk

picture_frame = CTkFrame(right_top, width=200, height=200, border_color="black", border_width=6, corner_radius=10, fg_color="transparent")
picture_frame.pack()
picture_frame.pack_propagate(False)  # Prevent frame from resizing

# Clickable label to display/upload image
picture_label = CTkLabel(
    picture_frame,
    text="Upload\n2x2 Picture",
    justify="center",
    font=("Arial", 12),
    bg_color="white",
    text_color="black"
)
picture_label.pack(fill="both", expand=True)
picture_label.bind("<Button-1>", lambda e: upload_image())

# Show personal details first
show_frame(personal_frame)

# Personal Information Title
personal_title_info =CTkLabel(scrollable_personal_frame, text= "PERSONAL INFORMATION", font= ("Id Inter", 25))
personal_title_info.pack(padx=20, pady=10, anchor=W)

# Personal Details Section
personal_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
personal_details.pack(padx=20, pady=10, fill="x")

# First Row
first_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
first_row_personal_details_frame.grid(row=0, column=0, columnspan=3, sticky="w", padx=20, pady=7)

surname_label = CTkLabel(first_row_personal_details_frame, text="Surname:", font=("Arial", 14), bg_color="transparent")
surname_label.grid(row=0, column=0, padx=20, pady=7, sticky="w")
surname_entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Surname", height= 35, fg_color= "transparent", bg_color= "transparent")
surname_entry.grid(row=1, column=0, padx=20, sticky="wn")

firstname_label = CTkLabel(first_row_personal_details_frame, text="Firstname:", font=("Arial", 14), bg_color="transparent")
firstname_label.grid(row=0, column=1, padx=20, pady=7, sticky="w")
firstname_entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Firstname", height= 35, fg_color= "transparent", bg_color= "transparent")
firstname_entry.grid(row=1, column=1, padx=20, sticky="wn")

middle_label = CTkLabel(first_row_personal_details_frame, text="Middle Initial:", font=("Arial", 14), bg_color="transparent")
middle_label.grid(row=0, column=2, padx=20, pady=7, sticky="w")
middle_entry = CTkEntry(first_row_personal_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Middle Initial", height= 35, fg_color= "transparent", bg_color= "transparent")
middle_entry.grid(row=1, column=2, padx=20, sticky="wn")

# Second Row
second_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
second_row_personal_details_frame.grid(row=1, column=0, columnspan=3, sticky="w", padx=20, pady=7)

gender_label = CTkLabel(second_row_personal_details_frame, text="Gender:", font=("Arial", 14), bg_color="transparent")
gender_label.grid(row=1, column=0, sticky="w", padx=40, pady=7)
gender_frame = CTkFrame(second_row_personal_details_frame, bg_color="transparent", fg_color= "transparent")
gender_frame.grid(row=2, column=0, sticky="w", padx=40)

male_var = IntVar()
female_var = IntVar()
none_binary_var = IntVar()

male_checkbox = CTkCheckBox(gender_frame, text="Male", variable=male_var)
male_checkbox.grid(row=0, column=0)
female_checkbox = CTkCheckBox(gender_frame, text="Female", variable=female_var)
female_checkbox.grid(row=0, column=1)
none_binary_checkbox = CTkCheckBox(gender_frame, text="Prefer not to answer", variable=none_binary_var)
none_binary_checkbox.grid(row=0, column=2)

age_label = CTkLabel(second_row_personal_details_frame, text="Age:", font=("Arial", 14), bg_color="transparent")
age_label.grid(row=1, column=1, sticky="w", padx=20, pady=7)
age_values = ["Select Age"] + [str(i) for i in range(1, 101)]
age_box = CTkComboBox(second_row_personal_details_frame, values=age_values, width=150, height= 35, bg_color= "transparent", state="normal")
age_box.set("Select Age")
age_box.grid(row=2, column=1, sticky="wn", padx=20)

birthdate_label = CTkLabel(second_row_personal_details_frame, text="Birthdate (MM/DD/YYYY):", font=("Arial", 14), bg_color="transparent")
birthdate_label.grid(row=1, column=2, sticky="w", padx=20, pady=7)
birthdate_frame = CTkFrame(second_row_personal_details_frame, bg_color="transparent", fg_color= "transparent")
birthdate_frame.grid(row=2, column=2, sticky="w", padx=20)

months = [
          "MM", "January", "February", "March", "April", 
          "May", "June", "July", "August", "September", 
          "October", "November", "December"
          ]
month_box = CTkComboBox(birthdate_frame, values=months, width=120, height= 35, bg_color= "transparent", state="normal")
month_box.set("MM")
month_box.grid(row=0, column=0, padx=3)

days = ["DD"] + [str(x) for x in range(1, 32)]
day_box = CTkComboBox(birthdate_frame, values=days, width=80, height= 35, bg_color= "transparent", state="normal")
day_box.set("DD")
day_box.grid(row=0, column=1, padx=3)

years = ["YYYY"] + [str(x) for x in range(1900, datetime.now().year + 11)]
year_box = CTkComboBox(birthdate_frame, values=years, width=150, height= 35, bg_color= "transparent", state="normal")
year_box.set("YYYY")
year_box.grid(row=0, column=2, padx=3)

# Third Row
third_row_personal_details_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
third_row_personal_details_frame.grid(row=3, column=0, columnspan=3, sticky="we", padx=20, pady=7)

birthplace_label = CTkLabel(third_row_personal_details_frame, text="Birthplace City:", font=("Arial", 14), bg_color="transparent")
birthplace_label.grid(row=0, column=0, sticky="w", padx=20, pady=7)
birthplace_entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Birthplace", height= 35, fg_color= "transparent", bg_color= "transparent")
birthplace_entry.grid(row=1, column=0, padx=20, sticky="wn")

nationality_label = CTkLabel(third_row_personal_details_frame, text="Nationality:", font=("Arial", 14), bg_color="transparent")
nationality_label.grid(row=0, column=1, sticky="w", padx=20, pady=7)
nationality_entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Nationality", height= 35, fg_color= "transparent", bg_color= "transparent")
nationality_entry.grid(row=1, column=1, padx=20, sticky="wn")

religion_label = CTkLabel(third_row_personal_details_frame, text="Religion:", font=("Arial", 14), bg_color="transparent")
religion_label.grid(row=0, column=2, sticky="w", padx=20, pady=7)
religions = [
             "Select Religion", "Roman Catholic", 
             "Iglesia ni Cristo", "Born Again", 
             "Protestant", "Muslim", "Other"
             ]
religion_box = CTkComboBox(third_row_personal_details_frame, values=religions, width=150, height= 35, bg_color= "transparent", state="readonly")
religion_box.set("Select Religion")
religion_box.grid(row=1, column=2, padx=20)

marital_status_label = CTkLabel(third_row_personal_details_frame, text="Marital Status:", font=("Arial", 14), bg_color="transparent")
marital_status_label.grid(row=0, column=3, sticky="w", padx=20, pady=7)
statuses = [
             "Select Marital Status", "Single", 
             "Married", "Widowed", "Separated"
            ]
marital_status_box = CTkComboBox(third_row_personal_details_frame, values=statuses, width=180, height= 35, bg_color= "transparent", state="readonly")
marital_status_box.set("Select Marital Status")
marital_status_box.grid(row=1, column=3, padx=20)

language_label = CTkLabel(third_row_personal_details_frame, text="Language Spoken:", font=("Arial", 14), bg_color="transparent")
language_label.grid(row=0, column=4, sticky="w", padx=20, pady=7)
language_entry = CTkEntry(third_row_personal_details_frame, width=200, font=("Arial", 14), placeholder_text="Language", height= 35, fg_color= "transparent", bg_color= "transparent")
language_entry.grid(row=1, column=4, padx=20, sticky="wn")

empty_frame = CTkFrame(personal_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=4, column=0, columnspan=3, sticky="we", padx=20, pady=7)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

# Address Title
address_title_info =CTkLabel(scrollable_personal_frame, text= "ADDRESS", font= ("Id Inter", 25))
address_title_info.pack(padx=20, pady=10, anchor=W)

# Address Details Section
address_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
address_details.pack(padx=20, pady=10, fill="x")

first_row_address_details_frame = CTkFrame(address_details, bg_color="transparent", fg_color= "transparent")
first_row_address_details_frame.grid(row=0, column=0, columnspan=3, sticky="we", padx=20, pady=7)

street_label = CTkLabel(first_row_address_details_frame, text="Street:", font=("Arial", 14), bg_color="transparent")
street_label.grid(row=0, column=0, padx=20, pady=7, sticky="w")
street_entry = CTkEntry(first_row_address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Street", height= 35, fg_color= "transparent", bg_color= "transparent")
street_entry.grid(row=1, column=0, padx=20, sticky="wn")

brgy_label = CTkLabel(first_row_address_details_frame, text="Barangay:", font=("Arial", 14), bg_color="transparent")
brgy_label.grid(row=0, column=1, padx=20, pady=7, sticky="w")
brgy_entry = CTkEntry(first_row_address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Barangay", height= 35, fg_color= "transparent", bg_color= "transparent")
brgy_entry.grid(row=1, column=1, padx=20, sticky="wn")

city_label = CTkLabel(first_row_address_details_frame, text="City/Municipality:", font=("Arial", 14), bg_color="transparent")
city_label.grid(row=0, column=2, padx=20, pady=7, sticky="w")
city_entry = CTkEntry(first_row_address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter City", height= 35, fg_color= "transparent", bg_color= "transparent")
city_entry.grid(row=1, column=2, padx=20, sticky="wn")

zip_code_label = CTkLabel(first_row_address_details_frame, text="Zip Code:", font=("Arial", 14), bg_color="transparent")
zip_code_label.grid(row=0, column=3, padx=20, pady=7, sticky="w")
zip_code_entry = CTkEntry(first_row_address_details_frame, width=170, font=("Arial", 14), placeholder_text="Enter Zip Code", height= 35, fg_color= "transparent", bg_color= "transparent")
zip_code_entry.grid(row=1, column=3, padx=20, sticky="wn")

province_label = CTkLabel(first_row_address_details_frame, text="Province:", font=("Arial", 14), bg_color="transparent")
province_label.grid(row=2, column=0, padx=20, pady=7, sticky="w")
province_entry = CTkEntry(first_row_address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Province", height= 35, fg_color= "transparent", bg_color= "transparent")
province_entry.grid(row=3, column=0, padx=20, sticky="wn")

country_label = CTkLabel(first_row_address_details_frame, text="Country:", font=("Arial", 14), bg_color="transparent")
country_label.grid(row=2, column=1, padx=20, pady=7, sticky="w")
country_entry = CTkEntry(first_row_address_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Country", height= 35, fg_color= "transparent", bg_color= "transparent")
country_entry.grid(row=3, column=1, padx=20, sticky="wn")

empty_frame = CTkFrame(address_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=1, column=0, columnspan=3, sticky="we", padx=20, pady=6)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

# Contact Information Title
contact_title_info =CTkLabel(scrollable_personal_frame, text= "CONTACT INFORMATION", font= ("Id Inter", 25))
contact_title_info.pack(padx=20, pady=10, anchor=W)

# Contact Information Details Section
contact_info_details = CTkFrame(scrollable_personal_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
contact_info_details.pack(padx=20, pady=10, fill="x")

first_row_contact_info_details_frame = CTkFrame(contact_info_details, bg_color="transparent", fg_color= "transparent")
first_row_contact_info_details_frame.grid(row=0, column=0, columnspan=3, sticky="we", padx=35, pady=7)

email_label = CTkLabel(first_row_contact_info_details_frame, text="Email:", font=("Arial", 14), bg_color="transparent")
email_label.grid(row=0, column=0, padx=10, pady=7, sticky="w")
email_entry = CTkEntry(first_row_contact_info_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Email", height= 35, fg_color= "transparent", bg_color= "transparent")
email_entry.grid(row=0, column=1, padx=10, pady=7, sticky="wn")
email_entry.bind("<FocusOut>", lambda event: validating_user_email())

num_label = CTkLabel(first_row_contact_info_details_frame, text="Contact Number:", font=("Arial", 14), bg_color="transparent")
num_label.grid(row=0, column=2, padx=10, pady=7, sticky="w")
num_entry = CTkEntry(first_row_contact_info_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
num_entry.grid(row=0, column=3, padx=10, pady=7, sticky="wn")

empty_frame = CTkFrame(contact_info_details, bg_color="transparent", fg_color= "transparent")
empty_frame.grid(row=1, column=0, columnspan=3, sticky="we", padx=20, pady=7)

empty_label = CTkLabel(empty_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)


# ==================== FAMILY BACKGROUND PAGE ====================
# Scrollable Frame
scrollable_family_background_frame = CTkScrollableFrame(family_frame)
scrollable_family_background_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
family_background_header = CTkFrame(scrollable_family_background_frame)
family_background_header.pack(pady=20)
CTkLabel(family_background_header, text="FAMILY BACKGROUND", font=("Arial", 28, "bold")).pack()

# Parent's Information Title
family_background_title_info =CTkLabel(scrollable_family_background_frame, text= "PARENT'S INFORMATION", font= ("Id Inter", 25))
family_background_title_info.pack(padx=20, pady=10, anchor=W)

# Parent's Details Section
parents_details = CTkFrame(scrollable_family_background_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
parents_details.pack(padx=20, pady=10, fill="x")

row_column_parents_details_frame = CTkFrame(parents_details, bg_color="transparent", fg_color= "transparent")
row_column_parents_details_frame.pack(anchor = CENTER, pady=30)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

father_label = CTkLabel(row_column_parents_details_frame, text="FATHER", font=("Arial", 14, "bold"), bg_color="transparent")
father_label.grid(row=0, column=1, padx=10, pady=7)

mother_label = CTkLabel(row_column_parents_details_frame, text="MOTHER", font=("Arial", 14, "bold"), bg_color="transparent")
mother_label.grid(row=0, column=2, padx=10, pady=7)

name_label = CTkLabel(row_column_parents_details_frame, text="Name:", font=("Arial", 14), bg_color="transparent")
name_label.grid(row=1, column=0, padx=10, pady=7, sticky="w")

name_father_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
name_father_entry.grid(row=1, column=1, padx=10, pady=7)

name_mother_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
name_mother_entry.grid(row=1, column=2, padx=10, pady=7)

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=2, column=0, padx=10, pady=7, sticky="w")

address_father_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_father_entry.grid(row=2, column=1, padx=10, pady=7)

address_mother_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_mother_entry.grid(row=2, column=2, padx=10, pady=7)

occupation_label = CTkLabel(row_column_parents_details_frame, text="Occupation:", font=("Arial", 14), bg_color="transparent")
occupation_label.grid(row=3, column=0, padx=10, pady=7, sticky="w")

occupation_father_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
occupation_father_entry.grid(row=3, column=1, padx=10, pady=7)

occupation_mother_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
occupation_mother_entry.grid(row=3, column=2, padx=10, pady=7)

phon_num_label = CTkLabel(row_column_parents_details_frame, text="Contact Number:", font=("Arial", 14), bg_color="transparent")
phon_num_label.grid(row=4, column=0, padx=10, pady=7, sticky="w")

phon_father_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
phon_father_entry.grid(row=4, column=1, padx=10, pady=7)

phon_mother_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Contact Number", height= 35, fg_color= "transparent", bg_color= "transparent")
phon_mother_entry.grid(row=4, column=2, padx=10, pady=7)

guardian_label = CTkLabel(row_column_parents_details_frame, text="GUARDIAN", font=("Arial", 14, "bold"), bg_color="transparent")
guardian_label.grid(row=5, column=0, padx=10, pady=7)

name_label = CTkLabel(row_column_parents_details_frame, text="Name:", font=("Arial", 14), bg_color="transparent")
name_label.grid(row=6, column=0, padx=10, pady=7, sticky="w")

name_guardian_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Name", height= 35, fg_color= "transparent", bg_color= "transparent")
name_guardian_entry.grid(row=6, column=1, padx=10, pady=7)

rel_label = CTkLabel(row_column_parents_details_frame, text="Relationship to Student:", font=("Arial", 14), bg_color="transparent")
rel_label.grid(row=6, column=2, padx=10, pady=7, sticky="s")

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=7, column=0, padx=10, pady=7, sticky="w")

address_guardian_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_guardian_entry.grid(row=7, column=1, padx=10, pady=7)

rel_guardian_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Relationship", height= 35, fg_color= "transparent", bg_color= "transparent")
rel_guardian_entry.grid(row=7, column=2, padx=10, pady=7)

occupation_label = CTkLabel(row_column_parents_details_frame, text="Occupation:", font=("Arial", 14), bg_color="transparent")
occupation_label.grid(row=8, column=0, padx=10, pady=7, sticky="w")

occupation_guardian_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Occupation", height= 35, fg_color= "transparent", bg_color= "transparent")
occupation_guardian_entry.grid(row=8, column=1, padx=10, pady=7)

# ==================== FAMILY BACKGROUND PAGE ====================
# Scrollable Frame
scrollable_educ_background_frame = CTkScrollableFrame(education_frame)
scrollable_educ_background_frame.place(x=0, y=0, relwidth=1, relheight=1)

# Headers
educ_background_header = CTkFrame(scrollable_educ_background_frame)
educ_background_header.pack(pady=20)
CTkLabel(educ_background_header, text="EDUCATIONAL BACKGROUND", font=("Arial", 28, "bold")).pack()

# Educ's Information Title
family_background_title_info =CTkLabel(scrollable_educ_background_frame, text= "EDUCATIONAL INFORMATION", font= ("Id Inter", 25))
family_background_title_info.pack(padx=20, pady=10, anchor=W)

# Educ's Details Section
educ_details = CTkFrame(scrollable_educ_background_frame, bg_color="transparent", fg_color= "transparent", border_width=6, corner_radius=10)
educ_details.pack(padx=20, pady=10, fill="x")

row_column_parents_details_frame = CTkFrame(educ_details, bg_color="transparent", fg_color= "transparent")
row_column_parents_details_frame.pack(anchor = CENTER, pady=30)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=0)

elem_label = CTkLabel(row_column_parents_details_frame, text="ELEMENTARY", font=("Arial", 14, "bold"), bg_color="transparent")
elem_label.grid(row=0, column=1, padx=10, pady=7)

schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
schl_label.grid(row=1, column=0, padx=10, pady=7, sticky="w")
schl_elem_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
schl_elem_entry.grid(row=1, column=1, padx=10, pady=7)

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=2, column=0, padx=10, pady=7, sticky="w")
address_elem_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_elem_entry.grid(row=2, column=1, padx=10, pady=7)

yr_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
yr_label.grid(row=3, column=0, padx=10, pady=7, sticky="w")
yr_elem_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
yr_elem_entry.grid(row=3, column=1, padx=10, pady=7)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=0, column=2)

js_label = CTkLabel(row_column_parents_details_frame, text="JUNIOR HIGH SCHOOL", font=("Arial", 14, "bold"), bg_color="transparent")
js_label.grid(row=0, column=3, padx=10, pady=7)

schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
schl_label.grid(row=1, column=2, padx=10, pady=7, sticky="w")
schl_js_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
schl_js_entry.grid(row=1, column=3, padx=10, pady=7)

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=2, column=2, padx=10, pady=7, sticky="w")
address_js_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_js_entry.grid(row=2, column=3, padx=10, pady=7)

yr_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
yr_label.grid(row=3, column=2, padx=10, pady=7, sticky="w")
yr_js_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
yr_js_entry.grid(row=3, column=3, padx=10, pady=7)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=4, column=0)

shs_label = CTkLabel(row_column_parents_details_frame, text="SENIOR HIGH SCHOOL", font=("Arial", 14, "bold"), bg_color="transparent")
shs_label.grid(row=4, column=1, padx=10, pady=7)

schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
schl_label.grid(row=5, column=0, padx=10, pady=7, sticky="w")
schl_shs_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
schl_shs_entry.grid(row=5, column=1, padx=10, pady=7)

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=6, column=0, padx=10, pady=7, sticky="w")
address_shs_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_shs_entry.grid(row=6, column=1, padx=10, pady=7)

yr_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
yr_label.grid(row=7, column=0, padx=10, pady=7, sticky="w")
yr_shs_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
yr_shs_entry.grid(row=7, column=1, padx=10, pady=7)

empty_label = CTkLabel(row_column_parents_details_frame, text="", font=("Arial", 14), bg_color="transparent")
empty_label.grid(row=4, column=2)

cg_label = CTkLabel(row_column_parents_details_frame, text="COLLEGE", font=("Arial", 14, "bold"), bg_color="transparent")
cg_label.grid(row=4, column=3, padx=10, pady=7, sticky="w")

trans_label = CTkLabel(row_column_parents_details_frame, text="For Transferees", font=("Arial", 14), bg_color="transparent")
trans_label.grid(row=4, column=3, padx=10, pady=7, sticky="e")

schl_label = CTkLabel(row_column_parents_details_frame, text="School Name:", font=("Arial", 14), bg_color="transparent")
schl_label.grid(row=5, column=2, padx=10, pady=7, sticky="w")
schl_cg_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter School", height= 35, fg_color= "transparent", bg_color= "transparent")
schl_cg_entry.grid(row=5, column=3, padx=10, pady=7)

address_label = CTkLabel(row_column_parents_details_frame, text="Address:", font=("Arial", 14), bg_color="transparent")
address_label.grid(row=6, column=2, padx=10, pady=7, sticky="w")
address_cg_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Address", height= 35, fg_color= "transparent", bg_color= "transparent")
address_cg_entry.grid(row=6, column=3, padx=10, pady=7)

yr_label = CTkLabel(row_column_parents_details_frame, text="Year Completed:", font=("Arial", 14), bg_color="transparent")
yr_label.grid(row=7, column=2, padx=10, pady=7, sticky="w")
yr_cg_entry = CTkEntry(row_column_parents_details_frame, width=250, font=("Arial", 14), placeholder_text="Enter Year Completed", height= 35, fg_color= "transparent", bg_color= "transparent")
yr_cg_entry.grid(row=7, column=3, padx=10, pady=7)

# ==================== Window Starter ====================
window.mainloop()