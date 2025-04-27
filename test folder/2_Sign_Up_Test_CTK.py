from PIL import Image
from customtkinter import *
from subprocess import call
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

window = CTk()
window.title("Login Portal")
window.geometry('1300x825')
window.resizable(False, False)

# ==================== Functions ====================

# Create and Save to Excel Function
def save_to_excel():
    if not data_validation_debugger():
        return
    
    username = user_entry.get()
    password = password_entry.get()

    try:
        wb = load_workbook("user_account_data.xlsx")
        if "Userdata" in wb.sheetnames:
            ws = wb["Userdata"]
        else:
            ws = wb.create_sheet("Userdata")

    except FileNotFoundError:

        wb = Workbook()
        ws = wb.active
        ws.title = "Userdata"
        ws.append(["Username", "Password"])

    ws.append([username, password])
    wb.save("user_account_data.xlsx")

    format_excel()
    show_data()

    messagebox.showinfo(title= "Success", message= "Account Saved")
    user_entry.delete(0, END)
    password_entry.delete(0, END)
    user_entry.insert(0, "Username")
    password_entry.insert(0, "Password")
     
# Format Fixer Function
def format_excel():
    wb = load_workbook("user_account_data.xlsx")
    ws = wb["Userdata"]

    # Bold Headers
    for cells in ws[1]:
        cells.font = Font(bold=True)

    # Auto Column Width
    for cols in ws.columns:
        max_length = max(len(str(cells.value)) for cell in cols)
        col_letter = get_column_letter(cols[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save("user_account_data.xlsx")

# New Window Data Shower Function
def show_data():
    wb = load_workbook("user_account_data.xlsx")
    ws = wb["Userdata"]

    data_window = CTkToplevel(window)
    data_window.title("Stored User Data")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            label = CTkLabel(data_window, text=value, borderwidth=1, relief="solid", padx=6, pady=3)
            label.grid(row=i, column=j)

# Input Validation & Debugger Function
def data_validation_debugger():
    username = user_entry.get()
    password = password_entry.get()

    if username and password and username != "Username" and password != "Password":
        print("\n\nSignup Form:\n\nUsername:", username, "\nPassword:", password)
        return True
    
    else:
        messagebox.showerror(title= "Error", message= "Please enter both Username and Password.")
        return False

# Username Delete & Restore Function
def on_username_click(event):
    if user_entry.get() == "Username":
        user_entry.delete(0, END)

def on_username_leave(event):
    name = user_entry.get()
    if name == "":
        user_entry.insert(0, "Username")

# Password Delete & Restore Function
def on_password_click(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, END)

def on_password_leave(event):
    password = password_entry.get()
    if password == "":
        password_entry.insert(0, "Password")

# Signup Window Switch Function
def go_login():
    window.destroy()
    call(["python", "1_Login_test.py"])
    pass

def handle_enter(event):
    if not save_to_excel():
        return

# //////////////////////////////////////////////////////////

# ==================== UI ====================
# Background & Banner Img
background_image = Image.open(".\wallhaven-85gxp2.png")
bg_img = CTkImage(light_image=background_image, dark_image=background_image, size=(1300, 825))
bg_label = CTkLabel(window, image=bg_img, text="")
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# //////////////////////////////////////////////////////////

# ==================== Frames ====================
frame = CTkFrame(window, border_width= 3, corner_radius= 15)
frame.place(relx = 0.5, rely = 0.5, anchor = CENTER)
# //////////////////////////////////////////////////////////

# ==================== Labels & Inputs ====================

# Project Name
project_label = CTkLabel(frame, text= "PROJECT UniPass", font= ("Times New Roman bold", 40))
project_label.grid(row=0, column=0, sticky="n", pady= 10, padx= 15)

# Short Description
short_desc_label =  CTkLabel(frame, text= "/Short Description/", font= ("Helvetica bold", 18))
short_desc_label.grid(row=1, column=0, sticky="n")

# Login
login_label = CTkLabel(frame, text= "Log In to Project", font= ("Helvetica bold", 17))
login_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)

# User Entry
user_entry = CTkEntry(frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Username")
user_entry.grid(row=3, column=0, sticky= "n", pady=15, padx= 15)
user_entry.bind("<FocusIn>", on_username_click)
user_entry.bind("<FocusOut>", on_username_leave)
user_entry.bind('<Return>', handle_enter)

# Divider Line
user_line = CTkFrame(frame, width=400, height=2, fg_color="white")
user_line.place(x=15, y=197)

# Password Entry
password_entry = CTkEntry(frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Password", show="*")
password_entry.grid(row=4, column=0, sticky= "n", pady=15, padx= 15)
password_entry.bind("<FocusIn>", on_password_click)
password_entry.bind("<FocusOut>", on_password_leave)
user_entry.bind('<Return>', handle_enter)


# Divider Line
pass_line = CTkFrame(frame, width=400, height=2, fg_color="white")
pass_line.place(x=15, y=255)

# //////////////////////////////////////////////////////////

# ==================== Buttons ====================

# Button Login
login_button = CTkButton(master=frame, text= "Sign Up", width=400, font= ("Arial bold", 15), command= save_to_excel)
login_button.grid(row=5, column=0, pady=15, padx= 15)

# Signup Text + Button
signup_frame = CTkFrame(frame)
signup_frame.grid(row=6, column=0, pady=20)

need_account_label = CTkLabel(signup_frame, text="Already a User?", font=("Arial", 12))
need_account_label.grid(row=0, column=0, padx=10)

signup_button = CTkButton(signup_frame, text="LOGIN", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2")
signup_button.grid(row=0, column=1, padx=5)

# //////////////////////////////////////////////////////////

window.mainloop()