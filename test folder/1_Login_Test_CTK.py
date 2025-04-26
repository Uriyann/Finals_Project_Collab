from PIL import Image
from customtkinter import *
from tkinter import messagebox
from openpyxl import load_workbook

window = CTk()
window.title("Login Portal")
window.geometry('1300x825')
window.resizable(False, False)

# ==================== Functions ====================

# Login Checker Function
def login_checker():
    if not data_validation_debugger():
        return

    username = user_entry.get()
    password = password_entry.get()

    try:
        wb = load_workbook("user_account_data.xlsx")
        ws = wb["Userdata"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                messagebox.showinfo("Login Success", f"Welcome, {username}!")
                return
            
        messagebox.showerror(title= "Login Failed", message= "Incorrect username or password.")

    except FileNotFoundError:
        messagebox.showerror(title= "Error", message= "User data file not found.")

# Input Validation & Debugger Function
def data_validation_debugger():

    username = user_entry.get()
    password = password_entry.get()

    if username and password and username != "Username" and password != "Password":
        print("\n\nData Entry Form:\n\n" \
              "Username:",username,
              "\nPassword:",password)
        return True
    
    else:
        messagebox.showerror(title= "Error", message= "Error. Input Required")
        return False

# Username Delete & Restore Function
def on_username_click(event):
    if user_entry.get() == "Username":
        user_entry.delete(0, 'end')

def on_username_leave(event):
    name = user_entry.get()
    if name == "":
        user_entry.insert(0, "Username")

# Password Delete & Restore Function
def on_password_click(event):
    if password_entry.get() == "Password":
        password_entry.delete(0, END)

def on_password_leave(event):
    password = password_entry.get()
    if password == "":
        password_entry.insert(0, "Password")

def handle_enter(event):
    if not login_checker():
        return
    pass

# //////////////////////////////////////////////////////////

# ==================== UI ====================
# Background & Banner Img
background_image = Image.open(".\wallhaven-85gxp2.png")
bg_img = CTkImage(light_image=background_image, dark_image=background_image, size=(1300, 825))
bg_label = CTkLabel(window, image=bg_img, text="")
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

side_image = Image.open(".\wallhaven-73616y.png")
side_img = CTkImage(light_image=side_image, dark_image=side_image, size=(550, 550))
side_label = CTkLabel(window, image=side_img, text="", corner_radius=10, fg_color="white")
side_label.grid(row=0, column=0, padx=90, pady=140)

# //////////////////////////////////////////////////////////

# ==================== Frames ====================
frame = CTkFrame(window, border_width= 3, corner_radius= 15)
frame.grid(row=0, column=1, padx=45)

# //////////////////////////////////////////////////////////

# ==================== Labels & Inputs ====================

# Project Name
project_label = CTkLabel(frame, text= "PROJECT UniPass", font= ("Times New Roman bold", 40))
project_label.grid(row=0, column=0, sticky="n", pady= 10, padx= 15)

# Short Description
short_desc_label =  CTkLabel(frame, text= "/Short Description/", font= ("Helvetica bold", 18))
short_desc_label.grid(row=1, column=0, sticky="n")

# Login
login_label = CTkLabel(frame, text= "Log In to Project", font= ("Microsoft YaHei UI Light", 17))
login_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)

# User Entry
user_entry = CTkEntry(frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Username")
user_entry.grid(row=3, column=0, sticky= "n", pady=15, padx= 15)
user_entry.bind("<FocusIn>", on_username_click)
user_entry.bind("<FocusOut>", on_username_leave)
user_entry.bind('<Return>', handle_enter)

# Divider Line
user_line = CTkFrame(frame, width=400, height=2, fg_color="white")
user_line.place(x=15, y=197)

# Password Entry
password_entry = CTkEntry(frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Password", show="*")
password_entry.grid(row=4, column=0, sticky= "n", pady=15, padx= 15)
password_entry.bind("<FocusIn>", on_password_click)
password_entry.bind("<FocusOut>", on_password_leave)
user_entry.bind('<Return>', handle_enter)


# Divider Line
pass_line = CTkFrame(frame, width=400, height=2, fg_color="white")
pass_line.place(x=15, y=255)

# //////////////////////////////////////////////////////////

# ==================== Buttons ====================

# Button Login
login_button = CTkButton(master=frame, text= "Log In", width=400, font= ("Arial bold", 15), command= login_checker)
login_button.grid(row=5, column=0, pady=15, padx= 15)

# Signup Text + Button
signup_frame = CTkFrame(frame)
signup_frame.grid(row=6, column=0, pady=20)

need_account_label = CTkLabel(signup_frame, text="Need an Account?", font=("Arial", 12))
need_account_label.grid(row=0, column=0, padx=10)

signup_button = CTkButton(signup_frame, text="SIGN UP", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2")
signup_button.grid(row=0, column=1, padx=5)

# //////////////////////////////////////////////////////////

window.mainloop()