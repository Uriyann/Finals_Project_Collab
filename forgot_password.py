from tkinter import *
from tkinter import messagebox
from customtkinter import *
import customtkinter as ctk
import openpyxl 
from openpyxl import load_workbook, Workbook
import tkinter as tk



def show_forg_pass():
    def reset_password():
        username = username_entry.get().strip()
        new_password = new_password_entry.get().strip()

        if not username or not new_password:
            messagebox.showerror("Input Error", "Both fields are required.")
            return

        if len(new_password) < 6:
            messagebox.showerror("Input Error", "Password must be at least 6 characters long.")
            return

        try:
            wb = load_workbook("user_account_data.xlsx")
            ws = wb["Userdata"]

            user_found = False
            for row in ws.iter_rows(min_row=2):
                if row[3].value == username:  # Column 4 (index 3) = Username
                    row[4].value = new_password  # Column 5 (index 4) = Password
                    user_found = True
                    break

            if user_found:
                wb.save("user_account_data.xlsx")
                messagebox.showinfo("Success", "Password updated successfully.")
                forg_pass_window.destroy()
            else:
                messagebox.showerror("User Not Found", "No user found with that username.")

        except FileNotFoundError:
            messagebox.showerror("Error", "User data file not found.")
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {str(e)}")

    forg_pass_window = CTkToplevel()
    forg_pass_window.title("Forgot Password")
    width, height = 400, 300
    x = (forg_pass_window.winfo_screenwidth() // 2) - (width // 2)
    y = (forg_pass_window.winfo_screenheight() // 2) - (height // 2)
    forg_pass_window.geometry(f"{width}x{height}+{x}+{y}")
    forg_pass_window.grab_set()

    CTkLabel(forg_pass_window, text="Enter your username:").pack(pady=(20, 5))
    username_entry = CTkEntry(forg_pass_window, width=250)
    username_entry.pack()

    CTkLabel(forg_pass_window, text="Enter new password:").pack(pady=(20, 5))
    new_password_entry = CTkEntry(forg_pass_window, width=250, show="*")
    new_password_entry.pack()

    CTkButton(forg_pass_window, text="Reset Password", command=reset_password).pack(pady=20)

    forg_pass_window.wait_window()