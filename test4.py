from tkinter import *
import openpyxl
from openpyxl import load_workbook, Workbook 
import tkinter as tk
from tkinter import messagebox

window = tk.Tk()  
window.title("Login Portal")
window.geometry('500x500')
window.resizable(False, False)  

def toggle_password_visibility():
    if  show_password_var.get() == 1:
            password_entry.configure(show="")
    else:
        password_entry.configure(show="*")

# def toggle_confirm_password_visibility():
#     if  show_confirm_password_var.get() == 1:
#             confirm_password_entry.configure(show="")
#     else:
#         confirm_password_entry.configure(show="*")



# # Create a new Excel workbook
def create_excel_file():
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Login"
    ws.append(["Name", "Email", "Username", "Password"])
    workbook.save("login.xlsx")
    print("Excel file created successfully!")

def edit_excel():

    nm = name_entry.get().strip()
    em = email_entry.get()
    un = Username_entry.get()
    pw = password_entry.get()
    wb = load_workbook("login.xlsx")
    ws = wb.active
    ws.append([nm, em, un, pw])
    wb.save("login.xlsx")
    messagebox.showinfo(title="Success",message= "Data saved successfully!")
    wb.save("login.xlsx")

try:
    wb = load_workbook("login.xlsx")
    if "studentdata" in wb.sheetnames:
        ws = wb["studentdata"]
    else:
        ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.title = "Login"
    ws.append(["Name", "email","username","password",])



def up_excel():
    
    nm = name_entry.get().strip()
    em = email_entry.get()
    un = Username_entry.get()
    pw = password_entry.get()

    wb = load_workbook("login.xlsx")
    ws = wb["studentdata"]



frame = tk. Frame(window, padx=20, pady=20)
frame.place(relx = 0.5, rely = 0.5, anchor = "center")


tk. Label (frame, text="Name", font=("Arial", 10, "bold")).grid(row=1, column=0)
tk. Label(frame, text="email",  font=("Arial", 10, "bold")).grid(row=3, column=0)
tk. Label(frame, text="username", font=("Arial", 10, "bold")).grid(row=5, column=0)
tk. Label(frame, text="password", font=("Arial", 10, "bold")).grid(row=7, column=0)


name_entry = tk. Entry(frame)
name_entry.grid(row=0, column=0)

email_entry= tk.Entry (frame)
email_entry.grid(row= 2, column=0)


Username_entry = tk. Entry(frame)
Username_entry.grid(row=4, column=0)


password_entry = tk. Entry(frame, show="*")
password_entry.grid(row=6, column=0)

show_password_var = tk.IntVar()
show_password_checkbox = tk.Checkbutton(frame, text="Show Password", variable=show_password_var, command=toggle_password_visibility)
show_password_checkbox.grid(row=8, column=0, sticky="w")    

# confirm_password_label = tk.Label(frame, text="Confirm Password:")
# confirm_password_label.grid (row=9, column=0)
# confirm_password_entry = tk.Entry(frame, show="*")
# confirm_password_entry.grid(row=9, column=1)

# show_confirm_password_var = tk.IntVar()
# show_confirm_password_checkbox = tk.Checkbutton(frame, text="Show Confirm Password", variable=show_confirm_password_var, command=toggle_confirm_password_visibility)
# show_confirm_password_checkbox.grid(row=9, column=2, sticky="w")

sub = tk.Button(frame, text="login", command=edit_excel)
sub.grid(row=10, column=0)
# up.grid(row=10, column=0, sticky="e")

create_excel_file()
frame.mainloop()
print("Excel file created successfully!")