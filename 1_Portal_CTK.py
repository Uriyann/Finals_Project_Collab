from PIL import Image, ImageDraw, ImageOps
from customtkinter import *
import customtkinter as ctk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re
import subprocess

# ==================== Window Setup ====================
window = CTk()
window.title("Login Portal")
height = 825
width = 1300
x = (window.winfo_screenwidth()//2)-(width//2) 
y = (window.winfo_screenheight()//2)-(height//2) 
window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
window.resizable(False, False)
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

mode = "dark"

def LOG_IN():
    # ==================== Functions ====================
    # Login Checker Function
    def login_checker():
        if not validate_login_name_pass_str():
            return

        username = login_user_entry.get()
        password = login_password_entry.get()

        try:
            wb = load_workbook("user_account_data.xlsx")
            ws = wb["Userdata"]

            if username == "admin" and password == "admin123":
                messagebox.showinfo("Login Success", "Welcome, Admin!")
                GO_TO_ADMIN_PANEL()
                return

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[3] == username and row[4] == password:
                    messagebox.showinfo("Login Success", f"Welcome, {username}!")
                    GO_TO_ENROLLMENT_FORM()
                    return
                
            messagebox.showerror(title= "Login Failed", message= "Incorrect username or password.")

        except FileNotFoundError:
            messagebox.showerror(title= "Error", message= "User data file not found.")

    def show_forg_pass():
        forg_pass_window = CTkToplevel(window)
        forg_pass_window.title("Forgot Password")
        height = 300
        width = 400
        x = (forg_pass_window.winfo_screenwidth()//2)-(width//2) 
        y = (forg_pass_window.winfo_screenheight()//2)-(height//2) 
        forg_pass_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
        forg_pass_window.grab_set()
        
        terms_label = CTkLabel(forg_pass_window, text="Forgot Password go here.")
        terms_label.pack(pady=20, padx=20)

        forg_pass_window.wait_window()

    # ==================== Debuggers ====================
    # Input Validation & Debugger Function
    def log_in_data_validation_debugger():

        username = login_user_entry.get()
        password = login_password_entry.get()

        if (username and password):
            print("\n\nData Entry Form:\n\n" \
                "Username:",username,
                "\nPassword:",password,
                "\n")
            return True
        
        else:
            messagebox.showerror(title= "Error", message= "Error. Input Required")
            return False

    # Str Debugger
    def validate_login_name_pass_str():
        if not log_in_data_validation_debugger():
            return

        name = login_user_entry.get()
        password = login_password_entry.get()
        if not re.fullmatch(r"[A-Za-z0-9_\s]+", name):      
            messagebox.showerror(title= "Invalid Input", message= "Special Characters are not allowed.")
            login_user_entry.delete(0, END)
            login_password_entry.delete(0, END)
            return False

        if len(password) < 6:
            messagebox.showerror(title= "Invalid Input", message= "Password must be at least 6 characters long.")
            login_password_entry.delete(0, END)
            return False
        
        return True

    # Portal Logo
    def make_rounded_image(image_path, size, corner_radius):
        
        image = Image.open(image_path).convert("RGBA")
        image = image.resize(size, Image.Resampling.LANCZOS)
        
        mask = Image.new("L", size, 0)
        draw = ImageDraw.Draw(mask)
        draw.rounded_rectangle((0, 0, size[0], size[1]), radius=corner_radius, fill=255)

        rounded_image = ImageOps.fit(image, size, centering=(0.5, 0.5))
        rounded_image.putalpha(mask)

        return rounded_image

    # ==================== Events ====================
    # Log In Enter Event
    def log_in_handle_enter(event):
        if not login_checker():
            return
        login_button.invoke()

    # Log In Light Switch Event
    def toggle_mode():
        global mode
        if mode == "dark":
            ctk.set_appearance_mode("light")
            mode = "light"
            toggle_button.configure(text="Switch to Dark Mode")
        else:
            ctk.set_appearance_mode("dark")
            mode = "dark"
            toggle_button.configure(text="Switch to Light Mode")

    # Log In Show Password Event
    def log_in_show_password():
        if show_password.get() == 1:
            login_password_entry.configure(show="")
        else:
            login_password_entry.configure(show="*")

    # Log In Window Switch Function
    def GO_TO_ENROLLMENT_FORM():
        window.destroy()
        subprocess.call(["python", "2_Enrollment_Form_CTK.py   "])

    # Log In Window Switch Function
    def GO_TO_ADMIN_PANEL():
        window.destroy()
        subprocess.call(["python", "4_Admin_Panel_CTK.py"])

    def GO_TO_LANDING_PAGE():
        window.destroy()
        subprocess.call(["python", "0_Landing_Page_TK.py"])

    # Signup Window Switch Function
    def SIGN_UP():
        # Deleting Login Frame and Creating Sign Up
        side_label.destroy()
        log_in_frame.destroy()
                
        # ==================== Functions ====================
        # Create and Save to Excel Function
        def save_to_excel():
            if not validate_signin_name_pass_str():
                return
            
            terms_accept = terms_accept_var.get()

            if terms_accept == "Accepted":
            
                first_name = sign_up_first_name_entry.get()
                last_name = sign_up_last_name_entry.get()
                email = sign_up_email_entry.get()
                username = sign_up_username_entry.get()
                password = sign_up_password_entry.get()

                try:
                    wb = load_workbook("user_account_data.xlsx")
                    if "Userdata" in wb.sheetnames:
                        ws = wb["Userdata"]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row[2] == email:
                                messagebox.showerror( title= "Error", message= "Email already exists. Please use a different email.")
                                return
                            elif row[3] == username:
                                messagebox.showerror( title= "Error", message= "Username already exists. Please use a different username.")
                                return
                            elif row[4] == password:
                                messagebox.showerror( title= "Error", message= "Password already exists. Please use a different password.")
                                return
                    else:
                        ws = wb.create_sheet("Userdata")

                except FileNotFoundError:

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Userdata"
                    ws.append(["First Name", "Last Name", "Email", "Username", "Password"])

                ws.append([first_name, last_name, email, username, password])
                wb.save("user_account_data.xlsx")

                format_excel()
                show_data()

                messagebox.showinfo(title= "Success", message= "Account Saved")
                sign_up_first_name_entry.delete(0, END)
                sign_up_last_name_entry.delete(0, END)
                sign_up_email_entry.delete(0, END)
                sign_up_username_entry.delete(0, END)
                sign_up_password_entry.delete(0, END)
                sign_up_confirm_password_entry.delete(0, END)

            else:
                messagebox.showerror(title="Error", message="You have not accepted the terms & conditions.")

        def show_terms_and_conditions():
            terms_window = CTkToplevel(window)
            terms_window.title("Terms and Conditions")
            terms_window.geometry("450x320")
            height = 320
            width = 450
            x = (terms_window.winfo_screenwidth()//2)-(width//2) 
            y = (terms_window.winfo_screenheight()//2)-(height//2) 
            terms_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            terms_window.resizable(False, False)
            terms_window.grab_set()

            terms_main_frame = CTkFrame(terms_window)
            terms_main_frame.pack(pady=10, padx=10)

            terms_text = """
                SYSTEM TERMS AND CONDITIONS

                1. Only authorized users are allowed to access this system.
                2. All data entered will be stored securely and used for official purposes only.
                3. Unauthorized modification or duplication of system data is prohibited.
                4. System activity may be logged for monitoring and audit purposes.
                5. By clicking "Accept", you agree to follow all guidelines stated.

                Please read carefully before proceeding.
                """
            
            terms_label = CTkLabel(terms_main_frame, text=terms_text, justify="left", wraplength=380, anchor="w", font=("Arial", 13))
            terms_label.pack(pady=3, padx=10)

            terms_accept_button = CTkButton(terms_main_frame, text="Accept", command=terms_window.destroy)
            terms_accept_button.pack(pady=5)

            terms_window.wait_window()

        def show_privacy_policy():
            priv_window = CTkToplevel(window)
            priv_window.title("Terms and Conditions")
            priv_window.geometry("450x320")
            height = 320
            width = 450
            x = (priv_window.winfo_screenwidth()//2)-(width//2) 
            y = (priv_window.winfo_screenheight()//2)-(height//2) 
            priv_window.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            priv_window.resizable(False, False)
            priv_window.grab_set()

            terms_main_frame = CTkFrame(priv_window)
            terms_main_frame.pack(pady=10, padx=10)

            terms_text = """
                SYSTEM PRIVACY POLICY

                1. Your personal data will be collected and processed in accordance with data protection laws.
                2. Information will not be shared with third parties without your consent.
                3. Data is stored securely and only authorized personnel can access it.
                4. You have the right to request access, correction, or deletion of your data.
                5. Continued use of the system indicates your agreement with our privacy practices.

                Please review carefully before proceeding.
                """
            
            terms_label = CTkLabel(terms_main_frame, text=terms_text, justify="left", wraplength=380, anchor="w", font=("Arial", 13))
            terms_label.pack(pady=3, padx=10)

            terms_accept_button = CTkButton(terms_main_frame, text="Accept", command=priv_window.destroy)
            terms_accept_button.pack(pady=5)

            priv_window.wait_window()

        # Format Fixer Function
        def format_excel():
            wb = load_workbook("user_account_data.xlsx")
            ws = wb["Userdata"]

            # Bold Headers
            for cells in ws[1]:
                cells.font = Font(bold=True)

            # Auto Column Width
            for cols in ws.columns:
                max_length = max(len(str(cells.value)) for cell in cols)
                col_letter = get_column_letter(cols[0].column)
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save("user_account_data.xlsx")

        # New Window Data Shower Function
        def show_data():
            wb = load_workbook("user_account_data.xlsx")
            ws = wb["Userdata"]

            data_window = CTkToplevel(window)
            data_window.title("Stored User Data")

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                for j, value in enumerate(row):
                    label = CTkLabel(data_window, text=value, width=1)
                    label.grid(row=i, column=j)

        # Input Validation & Debugger Function
        def sign_up_data_validation_debugger():
            first_name = sign_up_first_name_entry.get()
            last_name = sign_up_last_name_entry.get()
            email = sign_up_email_entry.get()
            username = sign_up_username_entry.get()
            password = sign_up_password_entry.get()
            confirm_password = sign_up_confirm_password_entry.get() 

            if (first_name and last_name and email and username and password and confirm_password):
                if password == confirm_password:
                    print("\n\nSignup Form:\n")
                    print("First Name:", first_name)
                    print("Last Name:", last_name)
                    print("Email:", email)
                    print("Username:", username)
                    print("Password:", password)
                    return True
                else:
                    messagebox.showerror(title="Error", message="Passwords do not match!")
                    return False
            else:
                messagebox.showerror(title="Error", message="Please fill out all fields.")
                return False
            
        # Str Debugger
        def validate_signin_name_pass_str():
            if not sign_up_data_validation_debugger():
                return

            firstname = sign_up_first_name_entry.get().strip()
            lastname = sign_up_last_name_entry.get().strip()
            email = sign_up_email_entry.get().strip()
            username = sign_up_username_entry.get().strip()
            passwords = sign_up_password_entry.get().strip()

            str_pattern = r"[A-Za-z0-9_\s]+"
            email_pattern = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"


            if not re.fullmatch(str_pattern, firstname):      
                messagebox.showerror(title= "Invalid Input", message= "Special Characters are not allowed.")
                return False
            
            if not re.fullmatch(str_pattern, lastname):      
                messagebox.showerror(title= "Invalid Input", message= "Special Characters are not allowed.")
                return False
            
            if not re.fullmatch(str_pattern, username):      
                messagebox.showerror(title= "Invalid Input", message= "Special Characters are not allowed.")
                return False
            
            if not re.fullmatch(email_pattern, email):      
                messagebox.showerror(title= "Invalid Input", message= "Invalid email format.")
                return False
            
            if len(passwords) < 6:
                messagebox.showerror(title= "Invalid Input", message= "Password must be at least 6 characters long.")
                return False
            
            return True
        
        # Sigh Up Light Switch Event
        def toggle_mode():
            global mode
            if mode == "dark":
                ctk.set_appearance_mode("light")
                mode = "light"
                toggle_button.configure(text="Switch to Dark Mode")
            else:
                ctk.set_appearance_mode("dark")
                mode = "dark"
                toggle_button.configure(text="Switch to Light Mode")

        # Message Confirmation Function
        def messagebox_confirmation():
            if messagebox.askyesno("Confirmation", "Are you sure you want to sign up?"):
                save_to_excel()
            else:
                return

        # Sign Up Enter Event
        def sign_up_handle_enter(event):
            if not save_to_excel():
                return
            sign_up_button.invoke()

        # Signup Window Switch Function
        def GO_BACK():
            # Deleting Sign Up and Calling Log In
            sign_up_frame.destroy()
            LOG_IN()

        # //////////////////////////////////////////////////////////

        # ==================== UI ====================
        # Background & Banner Img
        background_image = Image.open(r"C:\Users\M S I\Desktop\BSIT_Finals_Project_Collab\assets\wallhaven-85gxp2.png")
        bg_img = CTkImage(light_image=background_image, dark_image=background_image, size=(1300, 825))
        bg_label = CTkLabel(window, image=bg_img, text="")
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)

        # //////////////////////////////////////////////////////////

        # ==================== Frames ====================
        sign_up_frame = CTkFrame(window, border_width= 3, corner_radius= 15)
        sign_up_frame.place(relx = 0.5, rely = 0.5, anchor = CENTER)
        # //////////////////////////////////////////////////////////

        # ==================== Labels & Inputs ====================

        # Project Name
        sign_up_project_label = CTkLabel(sign_up_frame, text= "PROJECT UniPass", font= ("Times New Roman bold", 40))
        sign_up_project_label.grid(row=0, column=0, sticky="n", pady= 10, padx= 15)

        # Short Description
        short_desc_label =  CTkLabel(sign_up_frame, text= "/Secure and user-friendly student access portal./", font= ("Helvetica bold", 13))
        short_desc_label.grid(row=1, column=0, sticky="n")

        # Login
        sign_up_label = CTkLabel(sign_up_frame, text= "Sign Up to Project", font= ("Helvetica bold", 17))
        sign_up_label.grid(row=2, column=0, sticky="w", pady=15, padx= 15)
       
        # Firstname Entry
        sign_up_first_name_entry = CTkEntry(master=sign_up_frame, font= ("Arial", 16), border_width=0, width=195, placeholder_text="First name", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_first_name_entry.grid(row=3, column=0, sticky= "nw", padx= 15)
        sign_up_first_name_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=185, height=2)
        pass_line.grid(row=3, column=0, sticky= "ws", padx= 15)

        # Lastname Entry
        sign_up_last_name_entry = CTkEntry(sign_up_frame, font= ("Arial", 16), border_width=0, width=195, placeholder_text="Last name", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_last_name_entry.grid(row=3, column=0, sticky= "ne", padx= 15)
        sign_up_last_name_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=185, height=2)
        pass_line.grid(row=3, column=0, sticky= "es", padx= 27)

        # Email Entry
        sign_up_email_entry = CTkEntry(sign_up_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Email", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_email_entry.grid(row=4, column=0, sticky= "n", padx= 15)
        sign_up_email_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=400, height=2)
        pass_line.grid(row=4, column=0, sticky= S)

        # Username Entry
        sign_up_username_entry = CTkEntry(sign_up_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Username", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_username_entry.grid(row=5, column=0, sticky= "n", padx= 15)
        sign_up_username_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=400, height=2)
        pass_line.grid(row=5, column=0, sticky= S)

        # Password Entry
        sign_up_password_entry = CTkEntry(sign_up_frame, font= ("Arial", 16), border_width=0, width=195, placeholder_text="Password", show="*", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_password_entry.grid(row=6, column=0, sticky= "nw", padx= 15)
        sign_up_password_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=185, height=2)
        pass_line.grid(row=6, column=0, sticky= "ws", padx= 15)

        # Confirm Password Entry
        sign_up_confirm_password_entry = CTkEntry(sign_up_frame, font= ("Arial", 16), border_width=0, width=195, placeholder_text="Confirm Password", show="*", height= 55, fg_color= "transparent", bg_color= "transparent")
        sign_up_confirm_password_entry.grid(row=6, column=0, sticky= "ne", padx= 15)
        sign_up_confirm_password_entry.bind('<Return>', sign_up_handle_enter)

        # Divider Line
        pass_line = CTkFrame(sign_up_frame, width=185, height=2)
        pass_line.grid(row=6, column=0, sticky= "es", padx= 27)

        # //////////////////////////////////////////////////////////

        # ==================== Buttons ====================

        # Light Switch
        toggle_button = CTkButton(window, text="Swtich to Light Mode", command= toggle_mode, border_width= 3, corner_radius= 13, height=25)
        toggle_button.place(relx = 0.5, rely = 0.5, x= 495, y= -400)

        # Terms & Condition
        terms_accept_var = ctk.StringVar(value="Not Accepted")
        terms_chkbox = CTkCheckBox(sign_up_frame, text="I agree to the ", width=10, variable= terms_accept_var, onvalue= "Accepted", offvalue= "Not Accepted")
        terms_chkbox.grid(row=7, column=0, pady=6, padx= 15, sticky= "w")

        terms_button = CTkButton(sign_up_frame, text="terms & condition.", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2", command= show_terms_and_conditions, width=50, bg_color= "transparent")
        terms_button.grid(row=7, column=0, padx=119, sticky= "w")

        and_label = CTkLabel(sign_up_frame, text="and", font=("Arial", 12), bg_color="transparent")
        and_label.grid(row=8, column=0, padx=47, sticky= "w")

        priv_button = CTkButton(sign_up_frame, text="privacy policy.", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2", command= show_privacy_policy, width=50, bg_color= "transparent")
        priv_button.grid(row=8, column=0, padx=69, sticky= "w")

        # Button Login
        sign_up_button = CTkButton(sign_up_frame, text= "Sign Up", width=325, font= ("Arial bold", 15), command= save_to_excel, height= 35)
        sign_up_button.grid(row=9, column=0, pady=8, padx= 15)
        sign_up_button.configure(command=messagebox_confirmation)

        # Signup Text + Button
        log_frame = CTkFrame(sign_up_frame, fg_color= "transparent", bg_color= "transparent")
        log_frame.grid(row=10, column=0, pady=20)

        need_account_label = CTkLabel(log_frame, text="Already a User?", font=("Arial", 12))
        need_account_label.grid(row=0, column=0, padx=10)

        login_button = CTkButton(log_frame, text="LOGIN", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2", command= GO_BACK, width=50, bg_color= "transparent")
        login_button.grid(row=0, column=1, padx=5)

        # //////////////////////////////////////////////////////////

    # //////////////////////////////////////////////////////////

    # ==================== UI ====================
    # Background & Banner Img
    background_image = Image.open(r"C:\Users\M S I\Desktop\BSIT_Finals_Project_Collab\assets\wallhaven-85gxp2.png")
    bg_img = CTkImage(light_image=background_image, dark_image=background_image, size=(1300, 825))
    bg_label = CTkLabel(window, image=bg_img, text="")
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)

    side_image = Image.open(r"C:\Users\M S I\Desktop\BSIT_Finals_Project_Collab\assets\wallhaven-73616y.png")
    side_img = CTkImage(light_image=side_image, dark_image=side_image, size=(550, 550))
    side_label = CTkLabel(window, image=side_img, text="", corner_radius=10)
    side_label.place(relx = 0.5, rely = 0.5, x= -547, y= -275)

    # //////////////////////////////////////////////////////////

    back_homepge_btn = CTkButton(window, text="🏠", font=("Arial", 20, "bold"), command=GO_TO_LANDING_PAGE, width=50, height=50, corner_radius=10)
    back_homepge_btn.place(relx=0.5, rely=0.5, x= -635, y= -400)

    # ==================== Frames ====================
    log_in_frame = CTkFrame(window, border_width= 3, corner_radius= 15)
    log_in_frame.place(relx = 0.5, rely = 0.5, x= 90, y= -255)

    # //////////////////////////////////////////////////////////

    # ==================== Labels & Inputs ====================
    # Logo
    image_path = r"C:\Users\M S I\Desktop\BSIT_Finals_Project_Collab\assets\UniPass Logo.png"
    rounded_img = make_rounded_image(image_path, size=(70, 70), corner_radius=25)
    ctk_image = CTkImage(light_image=rounded_img, dark_image=rounded_img, size=(70, 70))
    label = CTkLabel(log_in_frame, image=ctk_image, text="")
    label.grid(row=0, column=0, padx=10, pady=10, sticky="n")

    # Project Name
    login_project_label = CTkLabel(log_in_frame, text= "PROJECT UniPass", font= ("Times New Roman bold", 40))
    login_project_label.grid(row=1, column=0, sticky="n", pady= 10, padx= 15)

    # Short Description
    short_desc_label =  CTkLabel(log_in_frame, text= "/Secure and user-friendly student access portal/", font= ("Helvetica bold", 13))
    short_desc_label.grid(row=2, column=0, sticky="n")

    # Login
    login_label = CTkLabel(log_in_frame, text= "Log In to Project", font= ("Helvetica bold", 17))
    login_label.grid(row=3, column=0, sticky="w", pady=15, padx= 15)

    # User Entry
    login_user_entry = CTkEntry(log_in_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Username or Email", height= 55, fg_color= "transparent", bg_color= "transparent", )
    login_user_entry.grid(row=4, column=0, sticky= "n", padx= 15)
    login_user_entry.bind('<Return>', log_in_handle_enter)

    # Divider Line
    user_line = CTkFrame(log_in_frame, width=400, height=2)
    user_line.grid(row=4, column=0, sticky= S)

    # Password Entry
    login_password_entry = CTkEntry(log_in_frame, font= ("Arial", 16), border_width=0, width=400, placeholder_text="Password", show="*", height= 55, fg_color= "transparent", bg_color= "transparent")
    login_password_entry.grid(row=5, column=0, sticky= "n", padx= 15)
    login_password_entry.bind('<Return>', log_in_handle_enter)

    # Divider Line
    pass_line = CTkFrame(log_in_frame, width=400, height=2)
    pass_line.grid(row=5, column=0, sticky= S)

    # //////////////////////////////////////////////////////////

    # ==================== Buttons ====================

    # Light Switch
    toggle_button = CTkButton(window, text="Switch to Light Mode", command= toggle_mode, border_width= 3, corner_radius= 13, height=25)
    toggle_button.place(relx = 0.5, rely = 0.5, x= 495, y= -400)

    # Show Password
    show_password = CTkCheckBox(log_in_frame, text="Show Password", width=10, command= log_in_show_password)
    show_password.grid(row=6, column=0, pady=6, padx= 15, sticky= "w")

    # Forgot Password
    forg_password = CTkButton(log_in_frame, text= "Forgot Password?", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="gray65", command= show_forg_pass, width=50, bg_color= "transparent")
    forg_password.grid(row=6, column=0, pady=6, padx= 15, sticky= "e")

    # Button Login
    login_button = CTkButton(log_in_frame, text= "Log In", width=325, font= ("Arial bold", 15), command= login_checker, height= 35, corner_radius= 16)
    login_button.grid(row=7, column=0, pady=8, padx= 15)

    # Signup Text + Button
    sign_frame = CTkFrame(log_in_frame, fg_color= "transparent", bg_color= "transparent")
    sign_frame.grid(row=8, column=0, pady=20)

    need_account_label = CTkLabel(sign_frame, text="Need an Account?", font=("Arial", 12))
    need_account_label.grid(row=0, column=0, padx=10)

    sign_up_button = CTkButton(sign_frame, text="SIGN UP", font=("Arial", 12), fg_color="transparent", hover_color="lightblue", text_color="dodgerblue2", command= SIGN_UP, width=50, bg_color= "transparent")
    sign_up_button.grid(row=0, column=1, padx=5)

    # //////////////////////////////////////////////////////////


# ==================== Window Starter ====================
LOG_IN()
window.mainloop()